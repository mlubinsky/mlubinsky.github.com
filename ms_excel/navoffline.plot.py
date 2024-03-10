import os
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
from openpyxl.styles import Color, Fill
metrics={}

metrics["Driving -Highway"] = [
"NonVDR_CEP50%(m)",
"NonVDR_CEP90%(m)",
"NonVDR_CEP95%(m)",
"NonVDR_CEP-STD(m)",
"NonVDR_CEP Max(m)",
"NonVDR_xTrack_Error50%(m)",
"NonVDR_xTrack_Error90%(m)",
"NonVDR_xTrack_Error95%(m)",
"NonVDR_xTrack_Error-STD(m)",
"NonVDR_xTrack_Error Max(m)",
"NonVDR_AlongTrack_Error50%(m)",
"NonVDR_AlongTrack_Error90%(m)",
"NonVDR_AlongTrack_Error95%(m)",
"NonVDR_AlongTrack_Error-STD(m)",
"NonVDR_AlongTrack_Error Max(m)",
"NonVDR_SOG_Error50%(m/s)",
"NonVDR_SOG_Error90%(m/s)",
"NonVDR_SOG_Error95%(m/s)",
"NonVDR_SOG_Error-STD(m/s)",
"NonVDR_SOG_Error Max(m/s)",
"NonVDR_Heading_Error50%(deg)",
"NonVDR_Heading_Error90%(deg)",
"NonVDR_Heading_Error95%(deg)",
"NonVDR_Heading_Error-STD(deg)",
"NonVDR_Heading_Error Max(deg)",
"NonVDR_EHPE_Ratio50%(m)",
"NonVDR_EHPE_Ratio90%(m)",
"NonVDR_EHPE_Ratio95%(m)",
"NonVDR_EHPE_Ratio-STD(m)",
"NonVDR_EHPE_Ratio Max(m)",
"NonVDR_Total_Fixes",
"NonVDR_Total_Fixes_Avg",
"Tunnel_out_TTFF",
"NonVDR_MX_static_sog_error_count_avg"
]

metrics["Driving -Highway VDR"]=[
"VDR_CEP50%(m)",
"VDR_CEP90%(m)",
"VDR_CEP95%(m)",
"VDR_CEP-STD(m)",
"VDR_CEP Max(m)",
"VDR_xTrack_Error50%(m)",
"VDR_xTrack_Error90%(m)",
"VDR_xTrack_Error95%(m)",
"VDR_xTrack_Error-STD(m)",
"VDR_xTrack_Error Max(m)",
"VDR_AlongTrack_Error50%(m)",
"VDR_AlongTrack_Error90%(m)",
"VDR_AlongTrack_Error95%(m)",
"VDR_AlongTrack_Error-STD(m)",
"VDR_AlongTrack_Error Max(m)",
"VDR_SOG_Error50%(m/s)",
"VDR_SOG_Error90%(m/s)",
"VDR_SOG_Error95%(m/s)",
"VDR_SOG_Error-STD(m/s)",
"VDR_SOG_Error Max(m/s)",
"VDR_Heading_Error50%(deg)",
"VDR_Heading_Error90%(deg)",
"VDR_Heading_Error95%(deg)",
"VDR_Heading_Error-STD(deg)",
"VDR_Heading_Error Max(deg)",
"VDR_EHPE_Ratio50%(m)",
"VDR_EHPE_Ratio90%(m)",
"VDR_EHPE_Ratio95%(m)",
"VDR_EHPE_Ratio-STD(m)",
"VDR_EHPE_Ratio Max(m)",
"VDR_Total_Fixes",
"VDR_Total_Fixes_Avg",
"VDR_Fixes_under_50m_avg",
"VDR_Fixes_over_50m_avg"
]

metrics["Driving - Parking lot in-out"] = [
"CEP50%(m)",
"CEP90%(m)",
"CEP95%(m)",
"CEP-STD(m)",
"CEP Max(m)",
"xTrack_Error50%(m)",
"xTrack_Error90%(m)",
"xTrack_Error95%(m)",
"xTrack_Error-STD(m)",
"xTrack_Error Max(m)",
"AlongTrack_Error50%(m)",
"AlongTrack_Error90%(m)",
"AlongTrack_Error95%(m)",
"AlongTrack_Error-STD(m)",
"AlongTrack_Error Max(m)",
"SOG_Error50%(m/s)",
"SOG_Error90%(m/s)",
"SOG_Error95%(m/s)",
"SOG_Error-STD(m/s)",
"SOG_Error Max(m/s)",
"Heading_Error50%(deg)",
"Heading_Error90%(deg)",
"Heading_Error95%(deg)",
"Heading_Error-STD(deg)",
"Heading_Error Max(deg)",
"EHPE_Ratio50%(m)",
"EHPE_Ratio90%(m)",
"EHPE_Ratio95%(m)",
"EHPE_Ratio-STD(m)",
"EHPE_Ratio Max(m)",
"Total_Fixes",
"Total_Fixes_Avg",
"NonVDR_Total_Fixes_Avg"
]

metrics["Driving - Urban Driving"] = [
"CEP50%(m)",
"CEP90%(m)",
"CEP95%(m)",
"CEP-STD(m)",
"CEP Max(m)",
"xTrack_Error50%(m)",
"xTrack_Error90%(m)",
"xTrack_Error95%(m)",
"xTrack_Error-STD(m)",
"xTrack_Error Max(m)",
"AlongTrack_Error50%(m)",
"AlongTrack_Error90%(m)",
"AlongTrack_Error95%(m)",
"AlongTrack_Error-STD(m)",
"AlongTrack_Error Max(m)",
"SOG_Error50%(m/s)",
"SOG_Error90%(m/s)",
"SOG_Error95%(m/s)",
"SOG_Error-STD(m/s)",
"SOG_Error Max(m/s)",
"Heading_Error50%(deg)",
"Heading_Error90%(deg)",
"Heading_Error95%(deg)",
"Heading_Error-STD(deg)",
"Heading_Error Max(deg)",
"EHPE_Ratio50%(m)",
"EHPE_Ratio90%(m)",
"EHPE_Ratio95%(m)",
"EHPE_Ratio-STD(m)",
"EHPE_Ratio Max(m)",
"Total_Fixes",
"Total_Fixes_Avg",
"MX_static_sog_error_count_avg"
]

metrics["Pedestrian - Urban"]=[
"CEP50%(m)",
"CEP90%(m)",
"CEP95%(m)",
"CEP-STD(m)",
"CEP Max(m)",
"xTrack_Error50%(m)",
"xTrack_Error90%(m)",
"xTrack_Error95%(m)",
"xTrack_Error-STD(m)",
"xTrack_Error Max(m)",
"AlongTrack_Error50%(m)",
"AlongTrack_Error90%(m)",
"AlongTrack_Error95%(m)",
"AlongTrack_Error-STD(m)",
"AlongTrack_Error Max(m)",
"SOG_Error50%(m/s)",
"SOG_Error90%(m/s)",
"SOG_Error95%(m/s)",
"SOG_Error-STD(m/s)",
"SOG_Error Max(m/s)",
"Heading_Error50%(deg)",
"Heading_Error90%(deg)",
"Heading_Error95%(deg)",
"Heading_Error-STD(deg)",
"Heading_Error Max(deg)",
"EHPE_Ratio50%(m)",
"EHPE_Ratio90%(m)",
"EHPE_Ratio95%(m)",
"EHPE_Ratio-STD(m)",
"EHPE_Ratio Max(m)",
"Total_Fixes",
"Total_Fixes_Avg"
]

metrics["Driving - Parking lot VDR (No signal)"] = [
"CEP50%(m)",
"CEP90%(m)",
"CEP95%(m)",
"CEP-STD(m)",
"CEP Max(m)",
"xTrack_Error50%(m)",
"xTrack_Error90%(m)",
"xTrack_Error95%(m)",
"xTrack_Error-STD(m)",
"xTrack_Error Max(m)",
"AlongTrack_Error50%(m)",
"AlongTrack_Error90%(m)",
"AlongTrack_Error95%(m)",
"AlongTrack_Error-STD(m)",
"AlongTrack_Error Max(m)",
"SOG_Error50%(m/s)",
"SOG_Error90%(m/s)",
"SOG_Error95%(m/s)",
"SOG_Error-STD(m/s)",
"SOG_Error Max(m/s)",
"Heading_Error50%(deg)",
"Heading_Error90%(deg)",
"Heading_Error95%(deg)",
"Heading_Error-STD(deg)",
"Heading_Error Max(deg)",
"EHPE_Ratio50%(m)",
"EHPE_Ratio90%(m)",
"EHPE_Ratio95%(m)",
"EHPE_Ratio-STD(m)",
"EHPE_Ratio Max(m)",
"Total_Fixes",
"Total_Fixes_Avg",
"VDR_Total_Fixes_Avg",
"Fixes_under_50m_avg"
]

#metrics["Driving - VDR-Sensor Hybrid (very weak signal)"] = [
metrics["Driving-VDR-Sensor Hybrid"] = [
"CEP50%(m)",
"CEP90%(m)",
"CEP95%(m)",
"CEP-STD(m)",
"CEP Max(m)",
"xTrack_Error50%(m)",
"xTrack_Error90%(m)",
"xTrack_Error95%(m)",
"xTrack_Error-STD(m)",
"xTrack_Error Max(m)",
"AlongTrack_Error50%(m)",
"AlongTrack_Error90%(m)",
"AlongTrack_Error95%(m)",
"AlongTrack_Error-STD(m)",
"AlongTrack_Error Max(m)",
"SOG_Error50%(m/s)",
"SOG_Error90%(m/s)",
"SOG_Error95%(m/s)",
"SOG_Error-STD(m/s)",
"SOG_Error Max(m/s)",
"Heading_Error50%(deg)",
"Heading_Error90%(deg)",
"Heading_Error95%(deg)",
"Heading_Error-STD(deg)",
"Heading_Error Max(deg)",
"EHPE_Ratio50%(m)",
"EHPE_Ratio90%(m)",
"EHPE_Ratio95%(m)",
"EHPE_Ratio-STD(m)",
"EHPE_Ratio Max(m)",
"Total_Fixes",
"Total_Fixes_Avg",
"Fixes_under_50m_avg",
"Fixes_over_50m_avg"
]

metrics["Tunnels"] = [
"PreTunnel_Fixes",
"PreTunnel_50%PosError",
"PreTunnel_95%PosError",
"PreTunnel_50%XTrackError",
"PreTunnel_95%XTrackError",
"PreTunnel_50%AlongError",
"PreTunnel_95%AlongError",
"PreTunnel_50%SpeedError",
"PreTunnel_95%SpeedError",
"PreTunnel_50%HeadingError",
"PreTunnel_95%HeadingError",
"InsideTunnel_Fixes",
"InsideTunnel_50%PosError",
"InsideTunnel_95%PosError",
"InsideTunnel_50%XTrackError",
"InsideTunnel_95%XTrackError",
"InsideTunnel_50%AlongError",
"InsideTunnel_95%AlongError",
"InsideTunnel_50%SpeedError",
"InsideTunnel_95%SpeedError",
"InsideTunnel_50%HeadingError",
"InsideTunnel_95%HeadingError",
"PostTunnel_Fixes",
"PostTunnel_50%PosError",
"PostTunnel_95%PosError",
"PostTunnel_50%XTrackError",
"PostTunnel_95%XTrackError",
"PostTunnel_50%AlongError",
"PostTunnel_95%AlongError",
"PostTunnel_50%SpeedError",
"PostTunnel_95%SpeedError",
"PostTunnel_50%HeadingError",
"PostTunnel_95%HeadingError"
]

# summary_sheet_metrics{}: key=sheet, value=(metric, formula)
# formula=1 means REF/DUT
# formula=0 means DUT/REF
summary_sheet_metrics={}

summary_sheet_metrics["Driving -Highway"]=[
    ("NonVDR_CEP90%(m)", 1),
    ("NonVDR_CEP95%(m)", 1),
    ("NonVDR_SOG_Error95%(m/s)", 1),
    ("NonVDR_MX_static_sog_error_count_avg",1),
    ("NonVDR_Heading_Error95%(deg)",1),
    ("Tunnel_out_TTFF",1)
]

summary_sheet_metrics["Driving -Highway VDR"]=[
    ("VDR_Total_Fixes_Avg", 0),
    ("VDR_Fixes_under_50m_avg", 0),
    ("VDR_Fixes_over_50m_avg/VDR_Total_Fixes_Avg", 3),    # Special case - ratio of metrics
    ("VDR_SOG_Error95%(m/s)", 1)
]

summary_sheet_metrics ["Driving - Parking lot in-out"]=[
    ("NonVDR_Total_Fixes_Avg", 0)
]

summary_sheet_metrics ["Driving - Urban Driving"]=[
    ("CEP95%(m)", 1),
    ("AlongTrack_Error95%(m)", 1),
    ("SOG_Error95%(m/s)", 1),
    ("MX_static_sog_error_count_avg", 1),
    ("Heading_Error95%(deg)", 1)
]

summary_sheet_metrics ["Pedestrian - Urban"]=[
    ("CEP95%(m)",1)
]

summary_sheet_metrics ["Driving - Parking lot VDR (No signal)"]=[
    ("VDR_Total_Fixes_Avg", 0),
    ("Fixes_under_50m_avg/VDR_Total_Fixes_Avg", 3), # Special case - ratio of metrics
    ("SOG_Error95%(m/s)", 1)
]

#summary_sheet_metrics ["Driving - VDR-Sensor Hybrid (very weak signal)"] = [
summary_sheet_metrics["Driving-VDR-Sensor Hybrid"] = [
    ("Fixes_under_50m_avg", 0),
    ("Fixes_over_50m_avg/Total_Fixes_Avg" , 3), # Special case - ratio of metrics
    ("SOG_Error95%(m/s)", 1)
]
##############################
def generate_excel_col_names():
##############################
    alphabet='ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    col_names=[]
    for l in alphabet:
        col_names.append(l)

    for l in alphabet[1:]:
     col_names.append('A'+l)

    return col_names

########################################
def parse_OriginalKPIStats(fname, metrics_of_interest,  reference="Kiwi"):
#####################################
    print("parsing", fname)

    metric_values=defaultdict(list)
    reference_values={} # key: metric name
    reference_index = None

    with open(fname, "r") as f:
        lines = f.readlines()
        n_metrics_found = 0

        for line_number, line in enumerate(lines):
            line_stripped = line.strip()
            columns = line_stripped.split(',')

            if line_number == 0:
               first_row = columns
               if reference not in first_row:
                  print("Reference not found", reference, "in file ", fname)
                  exit(1)
               else:
                  #number_of_columns =  len(columns)
                  #print("number_of_columns=", number_of_columns)
                  #print(columns)
                  reference_index = first_row.index(reference)
                  del  first_row[reference_index]

            else: # line_number > 0
               metric = columns[0]
               if metric in metrics_of_interest:

                  n_metrics_found +=1
                  if any(columns[1:]): # to exclude tiles like: PreTunnel_50%PosError,,,,,,,,,
                      if True: # columns[reference_index]: # reference value is not empty
                          metric_values[metric] = columns
                          reference_values[metric] = columns[reference_index]
                          del  metric_values[metric][reference_index] #  remove  reference_value
                      else:
                            print("Warning: empty reference value for", metric, "line=", line_number)
                  else:
                      print("Warning: excluding metric without data", metric, "line=", line_number)


        #print("n_metrics_found=", n_metrics_found, "len(metrics_of_interest)=", len(metrics_of_interest))

        if (n_metrics_found != len(metrics_of_interest)):
            for metric in metrics_of_interest:
                if metric not in metric_values:
                    print("Warning:", metric , "not found in", fname)

        l = len("NAVOFFLINE_") #eliminate this prefix from the build
        build_names=[] # list of build names
        for el in first_row:
            if el and el.startswith("NAVOFFLINE_"):
              el=el[l:]

            build_names.append(el)
        # Attention: 1st element in build names is None

    return build_names, metric_values, reference_values

###############################
def read_config_csv(config_file):
###############################
# Every line in file should have 2 comma-separated values: worksheet_name and the input file name
# Worksheet name should match the keys in metrics{} dictionary defined at the beginning of this file.
# Example:
#    Tunnels,      c:\CODE\SLOCATE-1226\Tunnels\OriginalKPIStats.csv
#    Driving -Highway,           c:\CODE\SLOCATE-1226\UC\OriginalKPIStats.csv

    f = open(config_file, 'r')
    lines = f.readlines()
    config = {}

    for line in lines:
      line = line.strip()
      if not line or line[0] == "#" or line.lower().startswith('worksheet'): # skip
        continue
      words = line.split(",")
      if len(words) != 2:
        print ("Error in line", line)
        print ("It should have 2 comm separated words: worksheet_name and file", line)
        exit(1)
      else:
         fname = words[1].strip()
         if not os.path.isfile(fname):
            print("no such file:", fname)
            exit(1)
         config[words[0].strip()] = fname
    return config

######################################################
def  plot(ws, metric_name, metric_number, n_columns ):
######################################################
            col_names = generate_excel_col_names()
            chart = LineChart()
            step=6
            vertical_shift = step * (metric_number)
            min_row = 3 + vertical_shift
            max_row = min_row + 2
            min_col=2 # because col=1 means column 'A'
            max_col = min_col + n_columns - 1
            data = Reference(worksheet=ws,
                 min_row=min_row,
                 max_row=max_row,
                 min_col=min_col,
                 max_col=max_col)

            if metric_number == 0:
                for col in range(min_col-1, max_col+1):
                    for row in range(min_row, max_row+1):
                        print("DEBUG", metric_name, 'col,row=', col,row,"value=", ws.cell(column=col,row=row).value)



            chart.add_data(data, from_rows=True, titles_from_data=False)
            chart.title = metric_name
            # chart.style = 13
            chart.x_axis.title="Build"
            if 'CEP' in metric_name:
                  chart.y_axis.title="m"
            elif ('Speed' in metric_name) or ('SOG' in metric_name):
                 chart.y_axis.title="m/s"
            elif 'Head' in metric_name:
                 chart.y_axis.title="deg"
            else:
                   chart.y_axis.title="?"

            if n_columns < len(col_names):
                  if metric_number % 2 == 0: #even
                      chart_column=col_names[n_columns+1]
                      chart_position = chart_column + str(min_row)

                  else: #odd
                      horizontal_shift=10 #TODO calculate it dynamically
                      chart_column=col_names[n_columns+1+horizontal_shift]
                      chart_position = chart_column + str(min_row-step)

            else:
                  print("ERROR", "n_columns=" , n_columns , "len(col_names)=", len(col_names))
                  exit(1)

            ws.add_chart(chart, chart_position)

#              anchor = TwoCellAnchor()
#              anchor._from.col = 0 #A
#              anchor._from.row = 8 # row 9, using 0-based indexing
#              anchor.to.col = 2 #C
#              anchor.to.row = 19 # row 20

#              chart.anchor = anchor

###############################
def apply_style(ws, reference):
###############################
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    pink_fill   = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

    font_bold = Font(bold=True)
    # font_bold.size = 18
    max_row_count = ws.max_row
    max_col_count = ws.max_column
    metric_name_line = False
    for row_num in range(1, max_row_count):
        value = ws.cell(row = row_num, column=1).value
        if not value and metric_name_line: # this is build line
            metric_name_line = False
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = green_fill
        if value and value.upper().startswith(("SPEED", "CEP", "HEAD", "PRE","POST","IN", "SOG")):
            ws.cell(row = row_num, column=1).font = font_bold
            metric_name_line = True
        elif value == 'TARGET':
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = yellow_fill
        elif value == 'GS24':
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = blue_fill
        elif value == reference:
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = pink_fill

############
def help():
###########
    print("Usage: required argument - config file name")
    print("  Every record in config file is comma-separated string")
    print("  The 1st column is the sheet name and the 2nd column is the path to OriginalKPIStats.csv")
    print("  The sheet name should be one of the following:")
    print()
    for key in metrics.keys():
        print(key)
    print()
    print("Optional 2nd argument: reference name (default: Kiwi)")
    exit(0)

##########################
def read_config_ini(fname):
##########################
    import configparser
    config = configparser.ConfigParser(interpolation=None)
    config.read(fname)
    print("all sections")
    print(config.sections())

    print("traverse sections one by one")
    for section in config.sections():
      print(section)
      print(config.items(section))
      print()

    return config

###################
def assign_colors(ws):
###################
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='7CFC00', end_color='7CFC00', fill_type='solid')
    red_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # row = 1 is header
    current_column_A = ws.cell(row = 2, column=1).value
    column_A_uniq_counter=0
    ## column_A_uniq_counter is used to interleave background color for even/odd sheets
    max_row_count = ws.max_row
    max_col_count = ws.max_column
    for row in range(2, max_row_count+1):
        column_A = ws.cell(row = row , column=1).value
        for col in range(1,max_col_count+1):
            if  column_A != current_column_A:
                current_column_A = column_A
                column_A_uniq_counter +=1

            val = ws.cell(row = row , column=col).value

            if  (col <= 3 and column_A_uniq_counter % 2 == 0):
                ws.cell(row = row, column = col).fill = blue_fill
            elif col > 3 and column_A_uniq_counter % 2 == 0 and not val:
                ws.cell(row = row, column = col).fill = blue_fill
            elif col > 3 and val:
                try:
                   val = float(val)
                   if val > 100.0:
                      color = green_fill
                   elif val > 90.0:
                      color = yellow_fill
                   else:
                     color = red_fill
                   ws.cell(row = row, column = col).fill = color
                except:
                    print("ERROR Cannot convert to float", val)

###################
def merge_cells(ws):
###################
    current_value=None
    max_row_count = ws.max_row
    for row_num in range(2, max_row_count):
        value = ws.cell(row = row_num, column=1).value
        if value != current_value:
            current_value=value
            #start_row = row_num
        else:
            ws.merge_cells(start_row=row_num-1, start_column=1, end_row=row_num, end_column=1)

############################################################
def create_summary_sheet(wb, build_names, summary_data):
############################################################
# summary_data is a dictionary: key=(sheet, metric, build) value=( DUT/REF or REF/DUT)
    print("inside summary sheet")

    for i, sheet in enumerate(wb.sheetnames):
        print(i, sheet)
    summary_sheet_name = "Performance Index"
    ws = wb.create_sheet(title=summary_sheet_name)

    header=["Sheet", "Metric"]
    # Add columns for build names
    header.extend(build_names)
    #print(header)

    placeholder = [None]*len(build_names)

    # Make header bold:
    font = Font(bold=True)
    for col, item in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = item
        cell.font = font


    rows=[]
    for i, sheet in enumerate(wb.sheetnames):
        if sheet == "Tunnels" or sheet == summary_sheet_name: # do not contribute to this sheet
            continue

        for metric in summary_sheet_metrics[sheet]:
            metric_name=metric[0]

            row=[sheet, metric_name] # because build names list start from None
            row.extend(placeholder)
            for build in build_names:
                key = (sheet, metric_name, build)
                if key in summary_data:
                    try:
                        build_index = header.index(build)
                        build_index +=1 #??

                        row[build_index] = summary_data[key]
                    except Exception as e:
                        print(str(e))
                else:
                    print("Warning: key not found in summary_data", key)

            rows.append(row)

    for row in rows:
        ws.append(row)

    assign_colors(ws)
    #merge_cells(ws)

##########################
def print_list(lst, name):
##########################
    print("printing list", name)
    for i, el in enumerate(lst):
            print(i,el)
    print()

###########################################
def process(config, output_file, reference):
###########################################
# creates all sheets
    print(output_file)
    print(config)
    wb = Workbook()
    ws = wb.active
    #print("ws.title=", ws.title)
    del wb['Sheet']

    longest_build_list=[]
    summary_data={}

    # Loop over sheets
    for sheet, fname in config.items():
        metrics_of_interest = metrics[sheet]
        ws = wb.create_sheet(title=sheet)
        print(sheet, "processing file:", fname)
        build_names, metrics_data, ref_data = parse_OriginalKPIStats(fname, metrics_of_interest, reference)
        ratio_metrics = find_ratio_metrics(sheet)

        #print("len(build_names)=", len(build_names),  "len(metrics_data)=", len(metrics_data),"len(ref_data)=", len(ref_data))

        if len(build_names) < 2 or len(metrics_data) == 0 or len(ref_data)  == 0 :
            print ("Error: after processing  " , fname )
            exit(0)

        if len(build_names) > len(longest_build_list):
            longest_build_list = build_names

        n_columns = len(build_names) # first element is None

        empty_record=[None] * n_columns

        metric_number = 0
        #Loop over metrics
        for metric_name, metric_vals in metrics_data.items():

            # row: metric name
            m = [metric_name]
            m.extend(empty_record[1:])
            ws.append(m)
            #  build names, 1st element  is None
            ws.append(build_names)

            record=["GS24"]
            ref_record=[reference]
            target_record=["TARGET"]

            #Loop over metric per build
            for i, v in enumerate(metric_vals[1:]):
                if v:
                    val = round(float(v),2)   # round to 2 digits after point

                record.append(val)

                if (ref_data[metric_name]):
                   ref_record.append(round(float(ref_data[metric_name]),2)) # round to 2 digits after point
                else:
                     ref_record.append(None)

                if sheet in summary_sheet_metrics:
                    if (metric_name, 0) in summary_sheet_metrics[sheet] : # here 0 means: REF/DUT
                        key = (sheet, metric_name, build_names[i])

                        if not ref_data[metric_name] or not v or float(v) == 0:
                            ratio=None
                        else:
                            ratio = 100 * float(ref_data[metric_name]) / float(v)
                            # round to 2 digits after point
                            ratio = round(ratio,2)

                        summary_data[key] = ratio
                        #summary_data[key] = formula
                    elif (metric_name, 1) in summary_sheet_metrics[sheet]: # here 1 means: DUT/REF
                        key = (sheet, metric_name, build_names[i])

                        if not ref_data[metric_name] or not v or float(ref_data[metric_name]) == 0:
                            ratio=None
                        else:
                            ratio = 100 * float(v) / float(ref_data[metric_name])
                            # round to 2 digits after point
                            ratio = round(ratio,2)

                        summary_data[key] = ratio

                    if metric_name in ratio_metrics:
                        metric2 = ratio_metrics[metric_name]
                        print("ratio metric", metric_name, metric2)


                target_record.append(0.00)

            ws.append(record)
            ws.append(ref_record)
            ws.append(target_record)
            ws.append(empty_record)

            apply_style(ws, reference)

            #plot(ws, metric_name, metric_number, n_columns) # TODO
            metric_number +=1

            #for pair in pairs:
            #    m1=pair[0]
            #    m2=pair[1]
            #    ratio_metric_name=m1+"/"+m2
            #    key = (sheet, ratio_metric_name, build_names[i])

    create_summary_sheet(wb, longest_build_list, summary_data)

    wb.save(output_file)
    wb.close()
    print()
    print("output_file:", output_file)

####################
def find_ratio_metrics(sheet):
###################
    pairs={}
    for ratio_of_metrics, val in summary_sheet_metrics[sheet]:
        if val == 3:  #REF/DUT
            print("found pair key=", ratio_of_metrics)
            m1m2 = ratio_of_metrics.split("/")
            pairs[m1m2[0]] = m1m2[1]

    return pairs


################################
def check_metrics_consistency():
################################
    print("check_metrics_consistency")

    for key, val in summary_sheet_metrics.items():

        if key not in metrics:
            print("Error: summary key=", key, " does not exists in metrics dictionary" )

        for  metric_and_formula  in val:
            metric_name = metric_and_formula[0]
            formula= metric_and_formula[1]
            if formula in [0,1] and metric_name not in metrics[key]:
                print("Error:", metric_name,   "does not exists under metrics[", key,"]")
            elif formula == 3:
                print("TODO", metric_name)

#################################
def check_config_keys(config):
################################
    for key, val in config.items():
        if key not in metrics:
            print("!!! Error: key ", key, "does not exists im metrics{}. metrics{} keys are:")

            for key, val in metrics.items():
               print (key)
            exit(1)

############################
if __name__ == "__main__":
############################
    check_metrics_consistency()

    n_args= len(sys.argv)
    if (n_args !=2 and n_args !=3):
     help()

    config_file = sys.argv[1]
    if not os.path.isfile(config_file):
            print("no such file:", config_file)
            exit(1)

    if n_args == 3:
       reference = sys.argv[2]
    else: # default reference
       reference="Kiwi"

    config = read_config_csv(config_file)
    check_config_keys(config)


    output_ms_excel_file=config_file+".xlsx"
    process(config, output_ms_excel_file, reference)

    exit(0)
