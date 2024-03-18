import os
import sys
from collections import defaultdict
import pandas as pd
#import csv
from openpyxl import Workbook
from openpyxl.styles import Font,  PatternFill

# summary_sheet_metrics{}: key=sheet, value=(metric, formula)
# formula=-1 means do not calculate ratio
# formula=0 means DUT/REF
# formula=1 means REF/DUT
# formula=3 means ratio of 2 metrics and REF/DUT

basic_metrics = [
    ("CEP50%(m)", 1),
    ("CEP95%(m)", 1),
    ("CEP Max(m)_Avg", 1),
    ("CEP Max(m)_Max", 1),

    ("SOG_Error50%(m/s)", 1),
    ("SOG_Error95%(m/s)", 1),
    ("SOG_Error Max(m/s)_Avg", 1),
    ("SOG_Error Max(m/s)_Max", 1),

    ("Heading_Error50%(deg)", 1),
    ("Heading_Error95%(deg)", 1),
    ("Heading_Error Max(deg)_Avg", 1),
    ("Heading_Error Max(deg)_Max", 1)
]
vdr_basic_metrics = [ ("VDR_" + s[0], s[1] ) for s in basic_metrics ]
nonvdr_basic_metrics = [ ("NonVDR_" + s[0], s[1] ) for s in basic_metrics ]

summary_metrics={}
mx_prefix={} # metrics to add MX_ prefix in output file

summary_metrics["Tunnels"]=[
("PreTunnel_Fixes", -1),
("PreTunnel_50%PosError", -1),
("PreTunnel_95%PosError", -1),
("PreTunnel_50%XTrackError", -1),
("PreTunnel_95%XTrackError", -1),
("PreTunnel_50%AlongError", -1),
("PreTunnel_95%AlongError", -1),
("PreTunnel_50%SpeedError", -1),
("PreTunnel_95%SpeedError", -1),
("PreTunnel_50%HeadingError", -1),
("PreTunnel_95%HeadingError", -1),
("InsideTunnel_Fixes_Count", -1),
("InsideTunnel_Fixes_Avg", -1),
("InsideTunnel_50%PosError", -1),
("InsideTunnel_95%PosError", -1),
("InsideTunnel_50%XTrackError", -1),
("InsideTunnel_95%XTrackError", -1),
("InsideTunnel_50%AlongError", -1),
("InsideTunnel_95%AlongError", -1),
("InsideTunnel_50%SpeedError", -1),
("InsideTunnel_95%SpeedError", -1),
("InsideTunnel_50%HeadingError", -1),
("InsideTunnel_95%HeadingError", -1),
("PostTunnel_Fixes", -1),
("PostTunnel_50%PosError", -1),
("PostTunnel_95%PosError", -1),
("PostTunnel_50%XTrackError", -1),
("PostTunnel_95%XTrackError", -1),
("PostTunnel_50%AlongError", -1),
("PostTunnel_95%AlongError", -1),
("PostTunnel_50%SpeedError", -1),
("PostTunnel_95%SpeedError", -1),
("PostTunnel_50%HeadingError", -1),
("PostTunnel_95%HeadingError", -1)
]

mx_prefix["Tunnels"]=[]

summary_metrics["Driving -Highway"] = nonvdr_basic_metrics.copy()

summary_metrics["Driving -Highway"] += [
    ("NonVDR_CEP90%(m)", 1),
    ("NonVDR_MX_static_sog_error_count",1),
    ("Tunnel_out_TTFF",1)
]

mx_prefix["Driving -Highway"]=[
    "NonVDR_CEP90%(m)",
    "NonVDR_CEP95%(m)"
]

summary_metrics["Driving -Highway VDR"] = vdr_basic_metrics.copy()
summary_metrics["Driving -Highway VDR"] += [
    ("VDR_Fixes_under_50m", 0),
    ("VDR_Total_Fixes", 0),     # also for using in ratio below
    ("VDR_Fixes_over_50m", -1), # just for using in ratio below
    ("VDR_Fixes_over_50m/VDR_Total_Fixes", 3),    # Special case - ratio of 2 metrics
]

mx_prefix["Driving -Highway VDR"]=[
  "VDR_Total_Fixes",
  "VDR_Fixes_under_50m",
  "VDR_Fixes_over_50m/VDR_Total_Fixes",
  "VDR_SOG_Error95%(m/s)"
]

summary_metrics ["Driving - Parking lot in-out"]=[
    ("NonVDR_Total_Fixes", 0)
]
mx_prefix["Driving - Parking lot in-out"]=[]

summary_metrics ["Driving - Urban Driving"] = basic_metrics.copy()
summary_metrics ["Driving - Urban Driving"] += [
    ("AlongTrack_Error95%(m)", 1),
    ("MX_static_sog_error_count", 1)
]

mx_prefix["Driving - Urban Driving"]=[
"CEP50%(m)",
"AlongTrack_Error95%(m)",
"SOG_Error95%(m/s)",
"Heading_Error95%(deg)"
]

summary_metrics ["Pedestrian - Urban"] = basic_metrics.copy()

mx_prefix["Pedestrian - Urban"]=[
   "CEP95%(m)"
]

# "Driving - Parking lot VDR (No signal)" - this name length > 31 chars, it cannot be used as MS Excel sheet name
summary_metrics ["Driving - Parking lot VDR"] = basic_metrics.copy()
summary_metrics ["Driving - Parking lot VDR"] += [
    ("Fixes_under_50m", 0),
    ("VDR_Total_Fixes", 0),  #   used in ratio below
    ("Fixes_over_50m", -1),   # just for using in ratio below
    ("Fixes_over_50m/VDR_Total_Fixes", 3)  # Special case - ratio of 2 metrics
]

mx_prefix["Driving - Parking lot VDR"]=[
   "VDR_Total_Fixes",
   "Fixes_over_50m/VDR_Total_Fixes",
   "SOG_Error95%(m/s)"
]

#"Driving - VDR-Sensor Hybrid (very weak signal)" - this name length > 31 chars, it cannot be used as MS Excel sheet name
summary_metrics ["Driving-VDR-Sensor Hybrid"] = basic_metrics.copy()
summary_metrics["Driving-VDR-Sensor Hybrid"] += [
    ("Fixes_under_50m", 0),
    ("Fixes_over_50m", -1),    # just for using in ratio below
    ("Total_Fixes", -1),        # just for using in ratio below
    ("Fixes_over_50m/Total_Fixes" , 3)  # Special case - ratio of 2 metrics
]
mx_prefix["Driving-VDR-Sensor Hybrid"] = [
    "Fixes_under_50m",
    "SOG_Error95%(m/s)",
    "Fixes_over_50m/Total_Fixes"
]

#############################################################################
def parse_OriginalKPIStats_csv(fname, metrics_of_interest, reference="Kiwi"):
#############################################################################
    df=pd.read_csv(fname, header=0 )
    if reference not in df.columns:
        print("Column ", reference, "not found in ", fname)
        exit(1)

    df = df.rename(columns={'Unnamed: 0': 'metric'})
    df['metric']=df['metric'].astype('string')
    filtered_df = df[df['metric'].isin(metrics_of_interest)]

    new_column_names = [col.replace('NAVOFFLINE_', '') for col in df.columns]
    df.columns = new_column_names
    return filtered_df


###############################
def read_config_csv(config_file):
###############################
# Every line in file should have 2 comma-separated values: worksheet_name and the input file name
# Worksheet name should match the keys in summary_metrics{} dictionary defined at the beginning of this file.
# Example:
#    Tunnels,           c:\CODE\SLOCATE-1226\Tunnels\OriginalKPIStats.csv
#    Driving -Highway,  c:\CODE\SLOCATE-1226\UC\OriginalKPIStats.csv

    f = open(config_file, 'r')
    lines = f.readlines()
    config = {}

    for line in lines:
      line = line.strip()
      if not line or line[0] == "#" or line.lower().startswith('worksheet'): # skip
        continue
      words = line.split(",")
      if len(words) != 2:
        print ("Error in line", line)
        print ("It should have 2 comma separated words: worksheet_name and file", line)
        exit(1)
      else:
          fname = words[1].strip()
          if not os.path.isfile(fname):
              print("no such file:", fname)
              exit(1)

          sheet = words[0].strip()

          if sheet not in summary_metrics:
            print("Error in config file: ", sheet, "does not exists im summary_metrics dictionary")
            print("Allowed keys are")
            print_list(summary_metrics.keys(), " ")
            exit(1)

          config[sheet] = fname
    return config

###############################
def assign_colors_to_sheet(ws, reference, metric_names, sheet):
###############################
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    pink_fill   = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    print("assign_colors_to_sheet metric_names=")
    print(metric_names)
    font_bold = Font(bold=True)
    # font_bold.size = 18
    max_row_count = ws.max_row
    max_col_count = ws.max_column
    metric_name_line = False
    for row_num in range(1, max_row_count+1):
        valueA = ws.cell(row = row_num, column=1).value  # value in 1st column A
        valueB = ws.cell(row = row_num, column=2).value  # value in 1st column A
        if row_num < 6:
            print(row_num, "A=", valueA, "B=", valueB)

        if not valueA and metric_name_line: # this is build line
            metric_name_line = False
            print("this is build line make it green")
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = green_fill

        if valueA and valueA in metric_names:
            ws.cell(row = row_num, column=1).font = font_bold
            if valueA in mx_prefix[sheet]:
                ws.cell(row = row_num, column=1).value = 'MX_' + valueA
            metric_name_line = True
        elif valueA == 'TARGET':
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = yellow_fill
        elif valueA == 'S24 Root' :
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = blue_fill
        elif valueA == reference:
            for col_num in range(2, max_col_count+1):
                ws.cell(row = row_num, column = col_num).fill = pink_fill

############
def help():
###########
    print("Usage: required argument - config file name")
    print("  Every record in config file is comma-separated string")
    print("  The 1st column is the sheet name and the 2nd column is the path to OriginalKPIStats.csv")
    print("  The sheet name should be one of the following:")
    print()
    for key in summary_metrics.keys():
        print(key)
    print()
    print("Optional 2nd argument: reference name (default: Kiwi)")
    exit(0)

#######################################
def assign_colors_to_summary_sheet(ws):
######################################
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='7CFC00', end_color='7CFC00', fill_type='solid')
    red_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # row = 1 is header, it contains A[1]: Sheet, B[1]:metric, C[1]: calc type (REF/DUT or DUT/REF), and all builds (ordered)
    current_column_A = ws.cell(row = 2, column=1).value
    column_A_uniq_counter=0
    ## column_A_uniq_counter is used to interleave background color for even/odd sheets
    max_row_count = ws.max_row
    max_col_count = ws.max_column
    for row in range(2, max_row_count+1):
        column_A = ws.cell(row = row , column=1).value # sheet name
        column_B = ws.cell(row = row , column=2).value # metric_name
        if column_A in mx_prefix and column_B in mx_prefix[column_A]:
            ws.cell(row = row , column=2).value  = 'MX_' + ws.cell(row = row , column=2).value
        for col in range(1, max_col_count+1):
            if  column_A != current_column_A:
                current_column_A = column_A
                column_A_uniq_counter +=1

            val = ws.cell(row = row , column=col).value

            if  (col <= 3 and column_A_uniq_counter % 2 == 0):
                ws.cell(row = row, column = col).fill = blue_fill
            elif col > 3 and column_A_uniq_counter % 2 == 0 and not val:
                ws.cell(row = row, column = col).fill = blue_fill
            elif col > 3 and val is not None:
                try:
                   val = float(val)
                   if val > 100.0:
                      color = green_fill
                   elif val > 90.0:
                      color = yellow_fill
                   else:
                     color = red_fill
                   ws.cell(row = row, column = col).fill = color
                except:
                    print("ERROR Cannot convert to float", val)

############################################################
def create_summary_sheet(wb, build_names, summary_data):
############################################################
# summary_data is a dictionary: key=(sheet, metric, build) value=( DUT/REF or REF/DUT)
    print("inside summary sheet")

    summary_sheet_name = "Performance Index"
    ws = wb.create_sheet(title=summary_sheet_name)

    header=["Sheet", "Metric"]
    # Add columns for build names
    header.extend(build_names)

    placeholder = [None]*len(build_names)

    # Make header bold:
    font = Font(bold=True)
    for col, item in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = item
        cell.font = font


    rows=[]
    for i, sheet in enumerate(wb.sheetnames):
        if sheet == "Tunnels" or sheet == summary_sheet_name: # do not contribute to summary sheet
            continue

        for metric in summary_metrics[sheet]:
            metric_name=metric[0]
            ratio_formula =  metric[1]
            if ratio_formula in [1,3]:
                ratio_name = 'REF/DUT'
            elif ratio_formula == 0:
                ratio_name = 'DUT/REF'
            elif ratio_formula == -1:
                continue
            else:
                print ("Unknown ratio_formula", ratio_formula)
                continue

            row=[sheet, metric_name, ratio_name] #   build names list start from None
            row.extend(placeholder[1:])
            for build in build_names[1:]:
                key = (sheet, metric_name, build)
                if key in summary_data:
                    try:
                        build_index = header.index(build)

                        row[build_index] = summary_data[key]
                    except Exception as e:
                        print(str(e))
                else:
                    print("Warning: key not found in summary_data", key)

            rows.append(row)

    for row in rows:
        ws.append(row)

    assign_colors_to_summary_sheet(ws)

##########################
def print_list(lst, name):
##########################
    print()
    print("printing list", name)
    for i, el in enumerate(lst):
            print(i,el)
    print()


###############################################################################
def report_missing_metrics(fname, metrics_of_interest, df):
###############################################################################
    for metric in metrics_of_interest:
        is_present = df["metric"].isin([metric])
        if not is_present.any() :
            if "/" in metric: # this is division of 2 metrics, it will be calculated in the code
                continue
            print ( metric, "  - missing in", fname)
            #print_list(df.columns, ' ')
            exit(0)



###########################################
def calc_metrics_ratio(df, composite_metrics):
###########################################
    # simple ratio is  DUT/REF or REF/DUT
    # composite ratio  is REF/DUT  where REF=( metric_1_ref/metric2_ref)  DUT = ( metric_1_dut/metric2/dut)
    print("start calc_metrics_ratio len(df)=", len(df))
    print(df["metric"])

    division_data=[]
    df_division=pd.DataFrame()

    for key, val in composite_metrics.items():
        print("key=",key)
        print ("val=",val)
        nom   = val["nominator"]
        denom = val["denominator"]
        nom_row =   df[df['metric'] == nom].reset_index(drop=True)
        denom_row = df[df['metric'] == denom].reset_index(drop=True)
        df_division = nom_row.iloc[:, 1:] / denom_row.iloc[:, 1:]
        division_data.append([key] + df_division.values.tolist()[0])

    #df_cols= [ col for col in df_division.columns if col !='formula' ]
    df_cols= [ col for col in df_division.columns]
    division_columns = ['metric'] + list(df_cols)
    df_division=pd.DataFrame(division_data, columns = division_columns)

    print("===df_division===")
    print(df_division)

    df_ref_by_ratio=pd.DataFrame()
    df_ref_by_ratio["metric"]=df_division["metric"]
    for col in df_division.columns:
        if col != 'metric' and col != reference: # and df_division.dtypes[col] = 'float64'
                df_ref_by_ratio[col] = df_division[reference] / df_division[col]


    print('df_ref_by_ratio=')
    print(df_ref_by_ratio)

    return df_division, df_ref_by_ratio


###########################################
def calc_ref_dut_ratio(df):
###########################################
    print("start calculate ratio")

    formulas=[0,1]   # REF/DUT or  1:DUR/REF
    df_ratio = df[df['formula'].isin(formulas)]
    #print("DEBUG len(df)=", len(df), "len(df_ratio)=", len(df_ratio))

    for i, row in df_ratio.iterrows():
        key=row['metric']
        for col in df.columns:
           if df[col].dtype =='float64' and col != reference and col != 'formula' :
              if row["formula"] == 1:
                  df_ratio.at[i,col] =  row[reference] / row[col]
              elif  row["formula"] == 0:
                 df_ratio.at[i,col] = row[col] / row[reference]
              elif  row["formula"] == -1:
                  pass
                 #print("DO nothing for row[ formula ] = ",row["formula"])
              else:
                  print("Error - unknown row[ formula ] = ",row["formula"])

    print("end calculate ratio")
    return df_ratio

#for column in df.columns:
#    if df[column].dtype =='float64' and column != reference:
#         ratio_df[column] = filtered_df[column] / df[reference]

# metric2formula=dict(basic_metrics)

##################################################################################################
def populate_worksheet(ws, sheet, df, df_ratio):
#################################################################################################
    #ws = wb.create_sheet(title=sheet)
    print ("START populate_worksheet()",sheet)
    #print("df_simple_ratio.columns=")
    #df_simple_ratio = df_simple_ratio.reset_index(drop=True)
    #print(df_simple_ratio.columns)
    #print("df_simple_ratio columns")
    builds=[]
    skip_len = len("NAVOFFLINE_")
    for i, col in enumerate(df_ratio.columns):
        print(i, col)
        if col.startswith("NAVOFFLINE_"):
            builds.append(col)
    builds_without_prefix = [ x[skip_len:] for x in builds]
    builds.sort()
    builds_without_prefix.sort()

    target_row=["TARGET"] + [100.00]*len(builds)
    ratio_row=[]

    # Report ratio metrics
    # for index, row in df_ratio.iterrows():
    for index, row in df.iterrows():
        metric = row["metric"]
        metric_name_row=[metric]
        #metric_name_row=[metric]
        #if metric in mx_prefix[sheet]:
        #    metric_name_row=["MX_"+metric]
        #else:
        #    metric_name_row=[metric]


        metrics_dict={}
        #df_metric_vals = df[df["metric"] ==  metric]
        #if len(df_metric_vals) !=1 :
        #    print("ERROR cannot find metric", metric)
        #    exit(0)
        #if len(df_metric_vals) !=1 :
        #    print("ERROR cannot find metric", metric)
        #    exit(0)
        # metrics_dict = df_metric_vals.squeeze()

        if row["formula"] in [0,1,3]:
            df_ratio_vals = df_ratio[df_ratio["metric"] ==  metric]
            if len(df_ratio_vals) !=1 :
                print("ERROR cannot find metric", metric)
                exit(0)
            metrics_dict = df_ratio_vals.squeeze()
            print("len(metrics_dict)=", len(metrics_dict))
            print(metrics_dict.keys())

        if row["formula"] == 1:
          ratio_row = [reference +"/S24 Root ratio"]
        elif  row["formula"] == 0:
          ratio_row = ["S24 Root/"+ reference + " ratio"]
        elif  row["formula"] == -1:
          ratio_row = None
        else:
          print("Warning unknown formula", row["formula"])


        metric_values_row=['S24 Root']
        reference_row=[reference]
        for b in builds:
              if row["formula"] in [0,1,3]:
                    # ratio_row.append(row[b])
                    ratio_row.append(metrics_dict[b])

              #metric_values_row.append(metrics_dict[b])
              metric_values_row.append(row[b])
              reference_row.append(row[reference])
              #reference_row.append(metrics_dict[reference])

        builds_row=[None] + builds_without_prefix

        ws.append(metric_name_row)
        ws.append(builds_row)
        
        ws.append(metric_values_row)
        ws.append(reference_row)

        #if row["formula"] in [0,1,3]:
        #  ws.append(ratio_row)

        ws.append(target_row)
        ws.append([])


##################################################################################################
def populate_worksheet_ratio(ws, sheet, df,  df_metrics_1_ratio, df_metrics_2_ratio):
#################################################################################################
    #ws = wb.create_sheet(title=sheet)
    print ("START populate_worksheet_ratio()",sheet)
    #print("df_simple_ratio.columns=")
    #df_simple_ratio = df_simple_ratio.reset_index(drop=True)
    #print(df_simple_ratio.columns)
    #print("df_simple_ratio columns")
    builds=[]
    skip_len = len("NAVOFFLINE_")
    for i, col in enumerate(df.columns):
        print(i, col)
        if col.startswith("NAVOFFLINE_"):
            builds.append(col)
    builds_without_prefix = [ x[skip_len:] for x in builds]
    builds.sort()
    builds_without_prefix.sort()

    target_row=["TARGET"] + [100.00]*len(builds)
    ratio_row=[]

    # Report ratio metrics
    for index, row in df.iterrows():
        metric = row["metric"]
        #metric_name_row=[metric]
        if metric in mx_prefix[sheet]:
            metric_name_row=["MX_"+metric]
        else:
            metric_name_row=[metric]


        df_metric_vals = df[df["metric"] ==  metric]
        if len(df_metric_vals) !=1 :
            print("ERROR cannot find metric", metric)
            exit(0)

        metrics_dict = df_metric_vals.squeeze()
        if row["formula"] == 1:
          ratio_row = [reference +"/S24 Root ratio"]
        elif  row["formula"] == 2:
          ratio_row = ["S24 Root/"+ reference + " ratio"]
        else:
          print("Warning unknown formula", row["formula"])
          exit(0)


        metric_values_row=['S24 Root']
        reference_row=[reference]
        for b in builds:
              ratio_row.append(row[b])
              metric_values_row.append(metrics_dict[b])
              #metric_values_row.append
              reference_row.append(row[reference])
              #reference_row.append(metrics_dict[reference])

        builds_row=[' '] + builds_without_prefix

        ws.append(metric_name_row)
        ws.append(builds_row)

        ws.append(metric_values_row)
        ws.append(reference_row)
        ws.append(ratio_row)
        ws.append(target_row)
        ws.append([])
#############################################################################
def populate_summary_sheet(wb,   all_builds, all_dfs):
#############################################################################
    print("START SUMMARY")
    builds = sorted(all_builds)
    builds_without_pefix=[ x[11:] for x in builds]
    print(builds)
    build_index = {item: index for index, item in enumerate(builds)}

    header=["Sheet","Metric"," "]
    header.extend(builds_without_pefix)
    ws = wb.create_sheet(title="Summary")
    ws.append(header)

    for sheet, df in all_dfs.items():
      print(df)
      print("summary sheet=", sheet, 'len(df)=', len(df))
      row=[sheet,'',''] # placeholders for metric and REF/DUT
      row +=['']*len(builds)
      print("len(row)=", len(row))

      for row_num, r in df.iterrows():
          #print(index, r["metric"])
          row[1] = r["metric"]
          formula = r['formula']
          if formula in [1,3]:
              row[2] = 'REF/DUT'
          elif formula == 0:
              row[2] = 'DUT/REF'
          else:
              row[2] = formula

          for b in builds:  # TODO builds need to constructed as union set of builds
             #if index + 3 > len(row) -1:
             #    print("out of index index=", index)
             #    continue
             val = r[b]
             if val:
                 val = round(100.0 * val, 2)

             row[build_index[b]+3]  = val


          ws.append(row)

      #ws.append([])

    assign_colors_to_summary_sheet(ws)


################################
def update_heading_columns(df):
###############################
    #float_cols = [col for col in df.columns if df[col].dtypes == np.float64]
    float_cols = df.select_dtypes(include=['float']).columns

    #-------------------------  
    def update_heading(row):
    #--------------------------
      if row['metric'].startswith('Heading'):
        for col in float_cols:
            row[col] = row[col] % 360
      return row

    df = df.apply(update_heading, axis=1)
    return df

###########################################
def process(config, output_file, reference):
###########################################
# creates all sheets
    print(output_file)
    # print(config)
    wb = Workbook()
    ws = wb.active
    #print("ws.title=", ws.title)
    del wb['Sheet']

    all_builds=set()
    #summary_data={}

    # Loop over sheets
    # ws_summary = wb.create_sheet(title="Summary")
    all_df_ratio={}
    for sheet, fname in config.items():

        metrics_with_formula=summary_metrics[sheet] # key: sheet; value: ( metric, formula) Here formula is: 1: REF/DUT  | 0:DUT/REF | 3:ratio of metrics REF/DUT = 3)

        metric2formula={}  # maps metric  to formula. The formula is: 1: REF/DUT  | 0:DUT/REF |  3:ratio of metrics REF/DUT     -1: do not calc ratio
        ratio_metrics={} # composite name with metric_1/metric_2  formula=3


        for m in metrics_with_formula:
            metric_name = m[0]
            formula = m[1]

            metric2formula[metric_name] = formula


            if formula == 3:  #3: REF/DUT
                nominator, denominator = metric_name.split("/")

                v={'nominator': nominator, 'denominator': denominator}
                ratio_metrics[metric_name]=v
                print("ratio_metrics", metric_name, v)

            if formula not in [-1,0,1,2,3]:
               print("Error unsupported formula", formula)
               exit(0)


        metrics_of_interest= metric2formula.keys()

        print(sheet, "processing file:", fname)


        df = parse_OriginalKPIStats_csv(fname, metrics_of_interest, reference)
        df = update_heading_columns(df)
        build_names=[]
        #skip_len = len("NAVOFFLINE_")
        for i, col in enumerate(df.columns):
          print(i, col)
          if col.startswith("NAVOFFLINE_"):
             build_names.append(col)
             all_builds.add(col)


        df["formula"] = df["metric"].map(metric2formula)
        report_missing_metrics(fname, metrics_of_interest, df)


        df_ratio = calc_ref_dut_ratio(df)
        print(df_ratio)
        ### Check: is formula = 3 present on df_ratio ?
        df_summary = df_ratio[df_ratio["formula"] == 3]
        print(df_summary)
        if len(df_summary) > 0:
          exit(0)

        all_df_ratio[sheet] = df_ratio
        #df_metrics_1_ratio, df_metrics_2_ratio = calc_metrics_ratio(df, ratio_metrics) TODO
        ws = wb.create_sheet(title=sheet)
        populate_worksheet(ws, sheet, df, df_ratio)

        #populate_worksheet_ratio(ws, sheet, df,   df_metrics_1_ratio, df_metrics_2_ratio)
        #continue

        #report_missing_metrics(fname, metrics_of_interest, metrics_data, ref_data)

        #if len(build_names) > len(longest_build_list):
        #    longest_build_list = build_names

        assign_colors_to_sheet(ws, reference, metrics_of_interest, sheet)
        break
        #populate_summary_sheet(ws_summary, sheet, longest_build_list, df_ratio)

    #create_summary_sheet(wb, longest_build_list, summary_data)
    populate_summary_sheet(wb,  all_builds, all_df_ratio)


    wb.save(output_file)
    wb.close()
    print()
    print("output_file:", output_file)

############################
if __name__ == "__main__":
############################

    n_args= len(sys.argv)

    #if (n_args !=2 and n_args !=3):
    #   help()

    #config_file = sys.argv[1]

    config_file = "navoffline_config.csv"
    print(config_file)
    if not os.path.isfile(config_file):
            print("no such file:", config_file)
            exit(1)

    if n_args == 3:
       reference = sys.argv[2]
    else: # default reference
       reference="Kiwi"


    config = read_config_csv(config_file)
    output_ms_excel_file= "absolut_new.xlsx"
    process(config, output_ms_excel_file, reference)

    exit(0)
