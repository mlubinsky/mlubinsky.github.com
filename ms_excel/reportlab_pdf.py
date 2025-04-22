import os
import sys
import argparse
import random
import logging
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import matplotlib.cm as cm
import seaborn as sns

from datetime import datetime, timedelta
from collections import defaultdict
from collections import Counter
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,  PatternFill, Alignment

from shutil import rmtree

C_SPOTACE_DIR=r"C:\SPOT-ACE"
print(C_SPOTACE_DIR)
sys.path.append(C_SPOTACE_DIR)

# Set up the logger
log_file="C:\\SPOT-ACE\\kpi_mx24\\mx_kpi_score_trend.log"
logging.basicConfig(
    level=logging.INFO,  # Log level
    format='%(asctime)s - %(message)s',  # Log format
    handlers=[
        logging.FileHandler(log_file),  # Log to a file
        logging.StreamHandler(sys.stdout)    # Log to the console
    ]
)

# Wrapper function for logging 
def log(message, level='info'):
    logging.info(message)

plt.set_loglevel('WARNING')

from utils.DBHelper import DBHelper
from spotemail import SpotEmailNew


#db = DBHelper(host='105.140.16.223')  # 'localhost', '105.140.16.228', '192.168.16.228'
#db = DBHelper(host='105.140.16.224')  # for testing
db = DBHelper(host='192.168.16.223') # prod


#OUTPUT_DIR=r"V:\RegressionLibrary\SPOT-ACE\kpi\Score"
OUTPUT_DIR=r"\\192.168.16.201\LocationResources\RegressionLibrary\SPOT-ACE\kpi\Score"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

IMG_DIR=os.path.join(SCRIPT_DIR, "images")
if os.path.exists(IMG_DIR):
       rmtree(IMG_DIR, ignore_errors=True)

DEBUG=False
THRESHOLD=90.0

PROJECT='MX'
DEFAULT_chipset='K43E1' 
DEFAULT_competitor_dut_models=str(('SM-S9260', 'SM-S928N', 'SM-S921U1', 'SM-S926U1', 'SM-S928U1', 'SM-S928B'))
DEFAULT_non_competitor_dut_models=str(('SM-S721B', 'SM-S721U', 'SM-S721N'))

highway=defaultdict(list)
highway["1. Horizontal Error"].append('Horizontal Error CDF 98%')
highway["1. Horizontal Error"].append('Along-Track Error CDF 98%')
highway["1. Horizontal Error"].append('AVG Horizontal Error Spec-out count')
highway["1. Horizontal Error"].append('Max Horizontal Error')

highway["2. Tunnel"].append('Max Tunnel Exit TTFF')
highway["2. Tunnel"].append('Max Tunnel Exit First Fix Horizontal Error')


highway["3. Speed"].append('Speed Error CDF 98% (km/h)')
highway["3. Speed"].append('AVG Speed Error Spec out count')
highway["3. Speed"].append('Max Speed Error (km/h)')


highway["4. VDR"].append('VDR - AVG Horizontal Error Spec(50m)-out count')
highway["4. VDR"].append('VDR - Max Horizontal Error')
highway["4. VDR"].append('VDR - Speed Error CDF 98% (km/h)')
highway["4. VDR"].append('VDR - AVG Speed Error Spec(20km/h)-out count')
highway["4. VDR"].append('VDR - Max Speed Error (km/h)')

highway["5. Stop"].append('AVG Static Speed Error count')
highway["5. Stop"].append('Max Static Speed Error Duration')
highway["5. Stop"].append('Max Static Speed Error (km/h)')

highway["6. Heading"].append('Heading Error CDF 98%')
highway["6. Heading"].append('Max Heading Error')

#  Urban Driving

urban_driving=defaultdict(list)
urban_driving["1. Horizontal Error"].append('Horizontal Error CDF 98%')
urban_driving["1. Horizontal Error"].append('Along-Track Error CDF 98%')
urban_driving["1. Horizontal Error"].append('AVG Horizontal Error Spec-out count')
urban_driving["1. Horizontal Error"].append('Max Horizontal Error')

urban_driving["2. Speed"].append('Speed Error CDF 98% (km/h)')
urban_driving["2. Speed"].append('AVG Speed Error Spec out count')
urban_driving["2. Speed"].append('Max Speed Error (km/h)')

urban_driving["3. Stop"].append('AVG Static Speed Error count')
urban_driving["3. Stop"].append('Max Static Speed Error Duration')
urban_driving["3. Stop"].append('Max Static Speed Error (km/h)')

urban_driving["4. Heading"].append('Heading Error CDF 98%')
urban_driving["4. Heading"].append('Max Heading Error')

#  Suwon Station Driving

suwon_driving=defaultdict(list)
suwon_driving["1. Horizontal Error"].append('Horizontal Error CDF 98%')
suwon_driving["1. Horizontal Error"].append('Along-Track Error CDF 98%')
suwon_driving["1. Horizontal Error"].append('AVG Horizontal Error Spec-out count')
suwon_driving["1. Horizontal Error"].append('Max Horizontal Error')

suwon_driving["2. Speed"].append('Speed Error CDF 98% (km/h)')
suwon_driving["2. Speed"].append('AVG Speed Error Spec out count')
suwon_driving["2. Speed"].append('Max Speed Error (km/h)')

suwon_driving["3. VDR"].append('VDR - AVG Horizontal Error Spec(50m)-out count')
suwon_driving["3. VDR"].append('VDR - Max Horizontal Error')
suwon_driving["3. VDR"].append('VDR - AVG Speed Error Spec(20km/h)-out count')
suwon_driving["3. VDR"].append('VDR - Max Speed Error (km/h)')

suwon_driving["4. Stop"].append('AVG Static Speed Error count')
suwon_driving["4. Stop"].append('Max Static Speed Error Duration')
suwon_driving["4. Stop"].append('Max Static Speed Error (km/h)')

suwon_driving["5. Heading"].append('Heading Error CDF 98%')
suwon_driving["5. Heading"].append('Max Heading Error')

######################################################
def populate_sheet(ws, df, date, df_dut_uniq):
#####################################################
# https://openpyxl.readthedocs.io/en/stable/styles.html
    print ("START populate_sheet() title =", ws.title, "date=", date)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    pink_fill   = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    grey_fill =  PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')

    font_bold = Font(bold=True, size=12)

   # make wide
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 45

    # Row 1
    first_row = list(df.columns)[0:2]
    for i, e in enumerate (list(df.columns)[2:]) :
        if e.endswith("score"):
            if e.startswith('REF'):
              first_row.append(None) 
            else: #TODO shouw build number here
                # find build  in df_dut_uniq
                dut = e[:-len('_score')]
                res = df_dut_uniq.loc[df_dut_uniq["dut_num"] == dut, "build_num"]
                if not res.empty:
                    val = res.values[0]
                else:
                    val = None

                first_row.append(val)

        else:
            first_row.append(e)

    ws.append(first_row)

    # Row 2: date, None, values and scores
    second_row=[str(date), None]
    for i, e in enumerate (list(df.columns)[2:]) :
        if e.endswith("score"):
            second_row.append('score')
        else:
            second_row.append('value')

    ws.append(second_row)

    # make 1st and 2nd rows color and bold
    for row_num in [1,2]:
      for col_num in range(1, ws.max_column+1):
        ws.cell(row = row_num, column = col_num).font = font_bold
        ws.cell(row = row_num, column = col_num).fill = grey_fill
    #TODO: consider using blue color for build number
    use_color = False
    for index, rec in df.iterrows():
        test = rec['Test']
        if index == 0:
            current_test = test
        row=[]

        if test != current_test:
            new_test = True
            use_color = not use_color
            current_test = test
        else:
            new_test = False

        for col in list(df.columns):
                row.append(rec[col])
        ws.append(row)
        #if use_color: alternate blue color
        #    for col in range(1, ws.max_column+1):
        #        ws.cell(row = index + 3, column = col).fill = blue_fill

        if rec['Criteria'] == 'Average':
            ws.cell(row = index + 3, column = 1).font = font_bold
            ws.cell(row = index + 3, column = 2).font = font_bold
            ws.cell(row = index + 3, column = 1).fill = green_fill
            ws.cell(row = index + 3, column = 2).fill = green_fill
            for col in range(3, ws.max_column+1):
                ws.cell(row = index + 3, column = col).font = font_bold
                v = ws.cell(row = index + 3, column = col).value
                if v and float(v) < THRESHOLD:
                    ws.cell(row = index + 3, column = col).fill = pink_fill
                else:
                    ws.cell(row = index + 3, column = col).fill = green_fill

#-------------------------------------------------------
def add_dut_columns(df, df_dut, is_ref, score_formula, sheet, date):
#-------------------------------------------------------
    # Adding to df the new columns from df_dut
    print("add_dut_columns() len(df) = ", len(df), ' len(df_dut)=', len(df_dut), " is_ref=", is_ref)

    # score_formula from list ['urban', 'suwon_driving', 'highway']

    df_score=pd.DataFrame()
    df_score['test']=pd.Series(dtype='string')
    df_score['criteria']=pd.Series(dtype='string')
    df_score['dut']=pd.Series(dtype='string')
    df_score['model']=pd.Series(dtype='string')
    df_score['build']=pd.Series(dtype='string')
    df_score['chipset']=pd.Series(dtype='string')
    df_score['val']=pd.Series(dtype='float32')
    df_score['score']=pd.Series(dtype='float32')

    dut_dict={} #key is dut_num; value is dictionary of {metric: value}

    for index, rec in df_dut.iterrows():
        build = rec['build_num']
        chipset=rec['chipset']
        dut_model=rec['dut_model']
        #print(build, chipset, ' is_ref=', is_ref)

        if is_ref:
            dut_num='REF '+ rec['dut_num']
        else:
            dut_num=rec['dut_num']

        if dut_num not in dut_dict:
            dut_dict[dut_num]={}
            dut_dict[dut_num+'_score']={}

        if rec['name'] == '_position2DError':
        #------------------------------------- 
            ##########
            # Tunnel
            ##########
            name="Max Tunnel Exit TTFF"
            test='Tunnel'
            #---------------------------
            val=rec['tunnel_out_ttffs_max']
            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                dut_dict[dut_num+'_score'][name] = 100 - val

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]

            name="Max Tunnel Exit First Fix Horizontal Error"
            test='Tunnel'
            #-----------------------------------------------
            val=rec['tunnel_out_ffposerr_max']
            if val is not None:
              dut_dict[dut_num][name] = round(val, 2)
              if rec['l5']:
                tr=30
              else:
                tr=50

              if val < tr:
                dut_dict[dut_num+'_score'][name] = 100
              else:
                dut_dict[dut_num+'_score'][name] = round(100 - (val - tr), 2)

            df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]

            #########
            #  Stop
            #########
            name='AVG Static Speed Error count'
            test='Stop'
            val=rec['static_speed_specouts']
            if val is not None and  str(val) != 'nan':
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                dut_dict[dut_num+'_score'][name] = round(100 - (val * 2))

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Max Static Speed Error Duration'
            test='Stop'
            val=rec['static_speed_error_duration_max']
            if val is not None  and  str(val) != 'nan':
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                dut_dict[dut_num+'_score'][name] = round(100 - (val * 3))

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Max Static Speed Error (km/h)'
            val=rec['static_speed_error_max']
            if val is not None and  str(val) != 'nan':
                val = round(val * 3.6, 2)
                dut_dict[dut_num][name] = val
                if val < 3.0:
                    dut_dict[dut_num+'_score'][name] = 100
                else:
                    dut_dict[dut_num+'_score'][name] = round(100 - ((val - 3) * 3))

                df_score.loc[len(df_score)]=[test, name, dut_num,dut_model,  build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Max Horizontal Error'
            test='Horizontal Error'
            if score_formula in ['urban', 'suwon_driving']:
                val=rec['max']
            else: # highway
                val=rec['non_tunnel_max']

            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                if rec['l5']:
                   tr=30
                else:
                    tr=50
                if val < tr:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = round(100 - (val - tr), 2)

            df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Horizontal Error CDF 98%'
            test='Horizontal Error'

            if score_formula in ['urban', 'suwon_driving']:
                val=rec['98']
                tr = 5
            else: # highway
                val=rec['non_tunnel_98']
                tr = 10

            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val

                if val < tr:
                    dut_dict[dut_num+'_score'][name] = 100
                else:
                    dut_dict[dut_num+'_score'][name] = round(100 - ((val - tr) * 2), 2)

            df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='AVG Horizontal Error Spec-out count'
            test='Horizontal Error'
            if score_formula in ['urban', 'suwon_driving']:
              if  rec['l5']:
                val=rec['poserrors_30']
              else:
                val=rec['poserrors_50']
            else: # highway
              if  rec['l5']:
                val=rec['non_tunnel_poserrors_30']
              else:
                val=rec['non_tunnel_poserrors_50']

            if val is not None:
              #name='AVG Horizontal Error Spec-out count'
              dut_dict[dut_num][name] = round(val, 2)
              if score_formula in ['urban','suwon_driving']:
                  dut_dict[dut_num+'_score'][name] = round(100 - (val*5), 2)
              else: # highway
                   dut_dict[dut_num+'_score'][name] = round(100 - val, 2)

            df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            val=rec['tunnel_poserrors_50']
            name = 'VDR - AVG Horizontal Error Spec(50m)-out count'
            test='VDR'
            if val is not None:
              dut_dict[dut_num][name] = round(val, 2)
              if val < 10:
                dut_dict[dut_num+'_score'][name] = 100
              else:
                dut_dict[dut_num+'_score'][name] = round(100 - (val - 10), 2)

              df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            val=rec['tunnel_speed_err_20km']
            test='VDR'
            if val is not None:
                name = 'VDR - AVG Speed Error Spec(20km/h)-out count'
                dut_dict[dut_num][name] = round(val, 2)
                dut_dict[dut_num+'_score'][name] = round(100 - (val*2), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            val=rec['tunnel_max']
            test='VDR'
            if val is not None:
                name = 'VDR - Max Horizontal Error'
                dut_dict[dut_num][name] =  round(val, 2)
                if val < 50:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = round(100 - (val - 50), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]

            #val=rec['speed_err_specout_cnt']
            val=rec['non_tunnel_speed_err_specout_cnt']
            test='Speed'
            if val is not None:
                name = 'AVG Speed Error Spec out count'
                dut_dict[dut_num][name] = round(val, 2)
                dut_dict[dut_num+'_score'][name] = round(100 - val, 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


        elif rec['name'] == '_alongTrackError':

            name='Along-Track Error CDF 98%'
            test='Horizontal Error'
            if score_formula in ['urban', 'suwon_driving']:
                val=rec['98']
            else:
                val=rec['non_tunnel_98']

            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                if val < 5:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = 100 - ((val - 5)*2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


        elif rec['name'] == '_headingError':
        #------------------------------------
            name='Max Heading Error'
            test='Heading'
            val=rec['hdg_error_max_exit30s']
            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                if score_formula in ['urban', 'suwon_driving']:
                   tr=5
                else:
                   tr=45
                if val < tr:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = round(100 - (val - tr), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Heading Error CDF 98%'
            test='Heading'
            if score_formula in ['urban', 'suwon_driving']:
                val=rec['98']
            else: # highway
                val=rec['non_tunnel_98']

            if val is not None:
                val = round(val, 2)
                dut_dict[dut_num][name] = val
                if  score_formula in ['urban', 'suwon_driving']:
                    if val < 3:
                        dut_dict[dut_num+'_score'][name] = round(100 - val, 2)
                    else:
                        dut_dict[dut_num+'_score'][name] = round(100 - ((val -3)*3), 2)
                else: # highway
                    dut_dict[dut_num+'_score'][name] = round(100 - (val * 2), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


        elif rec['name'] == 'SOG_Error':
        #--------------------------------
            name='Max Speed Error (km/h)'
            test='Speed'
            # val=rec['max']
            val=rec['non_tunnel_max']
            if val is not None:
                val = round(val*3.6, 2) # conversion from m/s to km/hr
                dut_dict[dut_num][name] = val
                if val < 10:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = round(100 - ((val - 10)*2), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='VDR - Max Speed Error (km/h)'
            test='VDR'
            val=rec['tunnel_max']
            if val is not None:
                val = round(val*3.6, 2) # conversion from m/s to km/hr
                dut_dict[dut_num][name] = val
                if val < 20:
                  dut_dict[dut_num+'_score'][name] = 100
                else:
                  dut_dict[dut_num+'_score'][name] = round(100 - ((val - 20)*2), 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='VDR - Speed Error CDF 98% (km/h)'
            test='VDR'
            val=rec['tunnel_98']
            if val is not None:
                val = round(val*3.6, 2) # conversion from m/s to km/hr
                dut_dict[dut_num][name] = val
                dut_dict[dut_num+'_score'][name] = 100 - val

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


            name='Speed Error CDF 98% (km/h)'
            test='Speed'
            if score_formula in ['urban', 'suwon_driving']:
                val=rec['98']
            else:
                val=rec['non_tunnel_98']
            if val is not None:
                val = round(val*3.6, 2) # conversion from m/s to km/hr
                dut_dict[dut_num][name] = val
                dut_dict[dut_num+'_score'][name] = round(100 - val*3, 2)

                df_score.loc[len(df_score)]=[test, name, dut_num, dut_model, build, chipset, val,  dut_dict[dut_num+'_score'][name]]


        else:
            #print("Skipping", rec['name'])
            continue

    # Fill NA with ?
    for dut_num, d in  dut_dict.items():
        #print ('Adding new column:',dut_num, ' len(d)=', len(d))
        df[dut_num]=df['Criteria'].map(d).fillna('?')


    df_score['date']=pd.to_datetime(date)
    df_score['test_loc']=sheet

    df_score['is_ref']=is_ref

    #print('END OF add_dut_columns() columns =')
    #print("--------------------------")
    #for col in df_score.columns:
    #            print(f"{col}: {df_score[col].dtype}")
    # - delete records from destination table with same date, test_loc and is_ref

    print('add_dut_columns() POPULATE_KPI_SCORE=', POPULATE_KPI_SCORE)
    if POPULATE_KPI_SCORE:
        sql=f"DELETE from kpi_score WHERE date='{date}' and test_loc='{sheet}' and is_ref={is_ref}"
        print(sql)
        db.dbExecute(sql)
        db.dbInsertDF(df_score, 'kpi_score')


    return df
#----------------------------------------
def combine_ref_and_dut(sheet, df_dut, df_competitors, date):
#----------------------------------------
    """
    Input: df_dut has fixed number of columns:
log_id: object
dut_model: object
dut_num: object
build_num: object
chipset: object
l5: object
name: object
50: object
95: object
98: object
max: object
vdr_98: object
vdr_max: object
tunnel_50: object
tunnel_95: object
tunnel_98: object
tunnel_max: object
non_tunnel_50: object
non_tunnel_95: object
non_tunnel_98: object
non_tunnel_max: object
poserrors_30: object
non_tunnel_poserrors_30: object
non_tunnel_poserrors_50: object
poserrors_50: object
tunnel_poserrors_50: object
tunnel_speed_err_20km: object
speed_err_specout_cnt: object
non_tunnel_speed_err_specout_cnt: object
tunnel_fixes: object
non_tunnel_fixes: object
fixes: object
static_speed_specouts: object
non_tunnel_static_speed_specouts: object
static_speed_error_duration_max: object
static_speed_error_max: object
hdg_error_max_exit30s: object
tunnel_out_ffposerr_max: object
tunnel_out_ttffs_max: object



    Returns: df with non-fixed number of columns: every dut is a separate column
    Test: object
    Criteria: object
    REF S24Ultra#2__SM-S928N: object
    REF S24Ultra#2__SM-S928N_score: object
    ...
    """
    print('combine_ref_and_dut(df_dut, df_competitors) sheet=', sheet)
    if df_dut.empty:
        print('df_dut.empty')
    #else:
    #    print('len(df_dut)=',len(df_dut))
    #    print(list(df_dut.columns))
    #    print('chipset=')
    #    print(df_dut['chipset'].tolist())
    #    print('build_num=')
    #    print(df_dut['build_num'].tolist())
    #    print(df_dut[['build_num','dut_num']])

    if df_competitors.empty:
        print('df_competitors.empty')
    #else:
    #    print('len(df_competitors)=',len(df_competitors))
    #    print(list(df_competitors.columns))
    #    print('competitors chipset=')
    #    print(df_competitors['chipset'].tolist())
    #    print('competitors build_num=')
    #    print(df_competitors['build_num'].tolist())
    #    print(df_competitors[['build_num','dut_num']])


    if sheet in ['Driving Highway ToSuwon','Driving Highway ToSeoul', 'Bundang Parking Garage']:
       score_formula='highway'
       current_dict = highway
    elif sheet == 'Bundang Urban Driving':
       score_formula='urban'
       current_dict = urban_driving
    elif sheet == 'Suwon Station':
       score_formula='suwon_driving'
       current_dict = suwon_driving
    else:
        print("Unknown sheet", sheet)
        df = pd.DataFrame()
        return df


    columns=['Test', 'Criteria']  # category, metric
    data=[]
    for category, metrics in current_dict.items():
      for index, metric in enumerate(metrics):

        data.append([str(category), str(metric)])

    df = pd.DataFrame(data, columns=columns)
    df['Test']=df['Test'].astype("string")
    df['Criteria']=df['Criteria'].astype("string")
    #print(df)


#len(df_dut)= 40
#['log_id', 'dut_model', 'dut_num', 'build_num', 'chipset', 'name', '50', '95', '98', 'max', 'tunnel_50', 'tunnel_95', 'tunnel_98', 'tunnel_max', 'non_tunnel_50', 'non_tunnel_95', 'non_tunnel_98', 'non_tunnel_max']

#len(df_competitors)= 4
#['log_id', 'dut_model', 'dut_num', 'chipset', 'name', '50', '95', '98', 'max', 'tunnel_50', 'tunnel_95', 'tunnel_98', 'tunnel_max', 'non_tunnel_50', 'non_tunnel_95', 'non_tunnel_98', 'non_tunnel_max']

# add  dut columns in dataframe for every dut
    if not df_competitors.empty:
        is_ref=True
        df = add_dut_columns(df, df_competitors, is_ref, score_formula, sheet, date )

    if not df_dut.empty:
        is_ref=False
        df = add_dut_columns(df, df_dut, is_ref, score_formula, sheet, date )


    # print(list(df.columns))
    #['Test', 'Criteria', 'REF S24Ultra#1__SM-S928N', 'REF S24Ultra#1__SM-S928N_score', 'REF S24Ultra#2__SM-S928N', 'REF S24Ultra#2__SM-S928N_score', 'REF S24Ultra#8__SM-S928N', 'REF S24Ultra#8__SM-S928N_score', 'S24FE#1', 'S24FE#1_score', 'S24FE#13', 'S24FE#13_score', 'S24FE#14', 'S24FE#14_score', 'S24FE#2', 'S24FE#2_score', 'S24FE#3', 'S24FE#3_score', 'S24FE#11', 'S24FE#11_score']

    return df


##########################
def get_ref_file_loc(date, file_location_patterns, competitor_dut_models): 
##########################

    if competitor_dut_models == 'ALL':
        competitor_filter=""
    else:
        competitor_filter=f" AND  dut_model in {competitor_dut_models} "

    sql_ref=f"SELECT  distinct file_loc \
       from gnslogindex \
       where competitor  {competitor_filter} \
       and date  = '{date}'"

    if DEBUG:
        print(sql_ref)

    results = db.dbFetchAll(sql_ref)

    ref_locations=[]
    for i, rec in enumerate(results):
      for p in file_location_patterns:
          if  p in rec[0]:
              ref_locations.append(rec[0])
              break

    return ref_locations

##########################
def get_dut_file_loc(date, file_location_patterns, chipset , non_competitor_dut_models ):
##########################
    print("get_dut_file_loc() date=", date)
    #   and file_loc in ({file_locations}) \

    if non_competitor_dut_models == 'ALL':
        non_competitor_filter=""
    else:
        non_competitor_filter=f" AND  dut_model in {non_competitor_dut_models} "

    if chipset == 'ALL':
        chipset_filter=""
    else:
        chipset_filter=f" AND  chipset='{chipset}' " 


    sql_ref=f"SELECT  distinct file_loc \
       from gnslogindex \
       where competitors_valid  {chipset_filter}  {non_competitor_filter} \
       and date  = '{date}'"

    #if DEBUG:
    print(sql_ref)
    results = db.dbFetchAll(sql_ref)

    dut_locations=[]
    for i, rec in enumerate(results):
        for  p in file_location_patterns:
          if  p in rec[0]:
              dut_locations.append(rec[0])
              break

    return dut_locations


####################################
def join_gnslogindex_duterror_navgps(date, file_locations, chipset , non_competitor_dut_models , only_builds=None, ignore_builds=None):
####################################
    print("join_gnslogindex_duterror_navgps() date=", date)

    build_filter=" "
    if only_builds:
        print(only_builds)
        build_filter = " AND g.build_num IN ({})".format(','.join(f"'{item}'" for item in only_builds))

    if ignore_builds:
        build_filter += " AND g.build_num NOT IN ({})".format(','.join(f"'{item}'" for item in ignore_builds))

    if non_competitor_dut_models == 'ALL':
        non_competitor_filter=""
    else:
        non_competitor_filter=f" AND g.dut_model in {non_competitor_dut_models} "

    if chipset == 'ALL':
        chipset_filter=""
    else:
        chipset_filter=f" AND  g.chipset='{chipset}' " 

    sql = f"SELECT g.log_id, g.dut_model, g.dut_num, g.build_num, g.chipset, g.l5, \
    e.name, \
    e.\"50\", e.\"95\" , e.\"98\", e.max, \
    e.vdr_98, e.vdr_max, \
    e.tunnel_50, e.tunnel_95, e.tunnel_98, e.tunnel_max, \
    e.non_tunnel_50 , \
    e.non_tunnel_95 , \
    e.non_tunnel_98 , \
    e.non_tunnel_max,  \
    n.poserrors_30,\
    n.non_tunnel_poserrors_30,\
    n.non_tunnel_poserrors_50,\
    n.poserrors_50,\
    n.tunnel_poserrors_50,\
    n.tunnel_speed_err_20km,\
    n.speed_err_specout_cnt, \
    n.non_tunnel_speed_err_specout_cnt, \
    n.tunnel_fixes,\
    n.non_tunnel_fixes,\
    n.fixes,\
    n.static_speed_specouts,\
    n.non_tunnel_static_speed_specouts, \
    n.static_speed_error_duration_max, \
    n.static_speed_error_max, \
    n.hdg_error_max_exit30s, \
    n.tunnel_out_ffposerr_max, \
    n.tunnel_out_ttffs_max \
    FROM gnslogindex g JOIN duterror e ON g.log_id = e.log_id \
    LEFT JOIN navgpssummary n ON n.log_id = g.log_id \
    WHERE g.competitors_valid \
    {chipset_filter} \
    {non_competitor_filter} \
    and g.competitors_valid \
    and g.file_loc in ({file_locations}) \
    and g.date  = '{date}' \
    {build_filter} \
    and e.name in ('_position2DError', 'SOG_Error', '_headingError', '_alongTrackError') "

    #if DEBUG:
    #print(sql)

    df = db.dbFetchAllDF(sql)
    #if DEBUG:
    print(" join_gnslogindex_duterror_navgps() df= ")
    #print(df)

    return df

####################################
def join_gnslogindex_duterror_navgps_competitors(date, file_locations, competitor_dut_models) :  
####################################
    # TODO - it is similar to function above: try to make 1 common function
    print("def join_gnslogindex_duterror_navgps_competitors() date=", date)
    if competitor_dut_models == 'ALL':
        competitor_filter=""
    else:
        competitor_filter=f" AND g.dut_model in {competitor_dut_models} "

    sql = f"SELECT g.log_id, g.dut_model, g.dut_num, g.chipset, g.l5, g.build_num, \
    e.name, \
    e.\"50\", e.\"95\" , e.\"98\", e.max, \
    e.vdr_98, e.vdr_max, \
    e.tunnel_50, e.tunnel_95, e.tunnel_98, e.tunnel_max, \
    e.non_tunnel_50 , \
    e.non_tunnel_95 as non_tunnel_95, \
    e.non_tunnel_98 as non_tunnel_98, \
    e.non_tunnel_max  as non_tunnel_max, \
    n.poserrors_30,\
    n.non_tunnel_poserrors_30,\
    n.non_tunnel_poserrors_50,\
    n.poserrors_50,\
    n.tunnel_poserrors_50,\
    n.tunnel_speed_err_20km,\
    n.speed_err_specout_cnt, \
    n.non_tunnel_speed_err_specout_cnt, \
    n.tunnel_fixes,\
    n.non_tunnel_fixes,\
    n.fixes,\
    n.static_speed_specouts,\
    n.non_tunnel_static_speed_specouts, \
    n.static_speed_error_duration_max, \
    n.static_speed_error_max, \
    n.hdg_error_max_exit30s, \
    n.tunnel_out_ffposerr_max, \
    n.tunnel_out_ttffs_max \
    FROM gnslogindex g JOIN duterror e ON g.log_id = e.log_id \
    LEFT JOIN navgpssummary n ON n.log_id = g.log_id \
    WHERE g.competitor \
    {competitor_filter} \
    and g.file_loc in ({file_locations}) \
    and g.date  = '{date}' \
    and e.name in ('_position2DError', 'SOG_Error', '_headingError', '_alongTrackError') "

    #if DEBUG:
    #  print(sql)

    df = db.dbFetchAllDF(sql)
    #if DEBUG:
    #    print("-- duterror_competitors dataframe --")
    #    print(df)
        #exit(0) # DEBUG
    return df

####################################
def avg_gnslogindex_duterror_navgps(date, file_locations,  chipset , non_competitor_dut_models , only_builds=None, ignore_builds=None):
####################################
    print("avg_gnslogindex_duterror_navgps() date=",date)

    build_filter=" "
    if only_builds:
        print(only_builds)
        build_filter = " AND g.build_num IN ({})".format(','.join(f"'{item}'" for item in only_builds))

    if ignore_builds:
        build_filter += " AND g.build_num NOT IN ({})".format(','.join(f"'{item}'" for item in ignore_builds))


    if non_competitor_dut_models == 'ALL':
        non_competitor_filter=""
    else:
        non_competitor_filter=f" AND g.dut_model in {non_competitor_dut_models} "

    if chipset == 'ALL':
        chipset_filter=""
    else:
        chipset_filter=f" AND  g.chipset='{chipset}' " 

    sql = f"SELECT g.dut_model, g.dut_num, g.build_num, g.chipset, g.l5,  \
    array_agg(g.log_id) as log_ids, \
    e.name, \
    AVG(e.\"50\") as \"50\", \
    AVG(e.\"95\") as \"95\", \
    AVG(e.\"98\") as \"98\", \
    AVG(e.vdr_98) as vdr_98, \
    AVG(e.vdr_max) as vdr_max, \
    MAX(e.max) as max,       \
    AVG(e.tunnel_50) as tunnel_50, \
    AVG(e.tunnel_95) as tunnel_95, \
    AVG(e.tunnel_98) as tunnel_98, \
    MAX(e.tunnel_max) as tunnel_max, \
    AVG(e.non_tunnel_50) as non_tunnel_50,   \
    AVG(e.non_tunnel_95) as non_tunnel_95,   \
    AVG(e.non_tunnel_98) as non_tunnel_98,   \
    MAX(e.non_tunnel_max) as non_tunnel_max, \
    AVG(n.poserrors_30) as poserrors_30,\
    AVG(n.poserrors_50) as poserrors_50,\
    AVG(n.non_tunnel_poserrors_30) as non_tunnel_poserrors_30,\
    AVG(n.non_tunnel_poserrors_50) as non_tunnel_poserrors_50,\
    AVG(n.tunnel_poserrors_50) as tunnel_poserrors_50,\
    AVG(n.tunnel_speed_err_20km) as tunnel_speed_err_20km,\
    AVG(n.speed_err_specout_cnt) as speed_err_specout_cnt, \
    AVG(n.non_tunnel_speed_err_specout_cnt) as non_tunnel_speed_err_specout_cnt, \
    AVG(n.tunnel_fixes) as tunnel_fixes,\
    AVG(n.non_tunnel_fixes) as non_tunnel_fixes,\
    AVG(n.fixes) as fixes,\
    AVG(n.static_speed_specouts) as static_speed_specouts,\
    AVG(n.non_tunnel_static_speed_specouts) as non_tunnel_static_speed_specouts, \
    MAX(n.static_speed_error_duration_max) as static_speed_error_duration_max, \
    MAX(n.static_speed_error_max) as static_speed_error_max, \
    MAX(n.hdg_error_max_exit30s) as hdg_error_max_exit30s, \
    MAX(n.tunnel_out_ffposerr_max) as tunnel_out_ffposerr_max, \
    MAX(n.tunnel_out_ttffs_max) as tunnel_out_ttffs_max \
    FROM gnslogindex g JOIN duterror e ON g.log_id = e.log_id \
    LEFT JOIN navgpssummary n ON n.log_id = g.log_id \
    WHERE g.competitors_valid \
    {chipset_filter} \
    {non_competitor_filter} \
    and g.file_loc in ({file_locations}) \
    and g.date  = '{date}' \
    {build_filter} \
    and e.name in ('_position2DError', 'SOG_Error', '_headingError', '_alongTrackError')  \
    group by  g.dut_model, g.dut_num, g.build_num, g.chipset, e.name, g.l5"


    #print(sql)
    df = db.dbFetchAllDF(sql)

    print("-- avg_gnslogindex_duterror_navgps() dataframe generated--")
    #print(df)


    return df

####################################
def avg_gnslogindex_duterror_navgps_competitors(date, file_locations, competitor_dut_models) :
####################################
    print("def avg_gnslogindex_duterror_navgps_competitors() date=", date)

    if competitor_dut_models == 'ALL':
        competitor_filter=""
    else:
        competitor_filter=f" AND g.dut_model in {competitor_dut_models} "

    sql = f"SELECT g.dut_model, g.dut_num, g.chipset, g.l5, g.build_num,  \
    array_agg(g.log_id) as log_ids, \
    e.name, \
    AVG(e.\"50\") as \"50\", \
    AVG(e.\"95\") as \"95\", \
    AVG(e.\"98\") as \"98\", \
    AVG(e.vdr_98) as vdr_98, \
    AVG(e.vdr_max) as vdr_max, \
    MAX(e.max) as max,       \
    AVG(e.tunnel_50) as tunnel_50, \
    AVG(e.tunnel_95) as tunnel_95, \
    AVG(e.tunnel_98) as tunnel_98, \
    MAX(e.tunnel_max) as tunnel_max, \
    AVG(e.non_tunnel_50) as non_tunnel_50,   \
    AVG(e.non_tunnel_95) as non_tunnel_95,   \
    AVG(e.non_tunnel_98) as non_tunnel_98,   \
    MAX(e.non_tunnel_max) as non_tunnel_max, \
    AVG(n.poserrors_30) as poserrors_30,\
    AVG(n.poserrors_50) as poserrors_50,\
    AVG(n.non_tunnel_poserrors_30) as non_tunnel_poserrors_30,\
    AVG(n.non_tunnel_poserrors_50) as non_tunnel_poserrors_50,\
    AVG(n.tunnel_poserrors_50) as tunnel_poserrors_50,\
    AVG(n.tunnel_speed_err_20km) as tunnel_speed_err_20km,\
    AVG(n.speed_err_specout_cnt) as speed_err_specout_cnt, \
    AVG(n.non_tunnel_speed_err_specout_cnt) as non_tunnel_speed_err_specout_cnt, \
    AVG(n.tunnel_fixes) as tunnel_fixes,\
    AVG(n.non_tunnel_fixes) as non_tunnel_fixes,\
    AVG(n.fixes) as fixes,\
    AVG(n.static_speed_specouts) as static_speed_specouts,\
    AVG(n.non_tunnel_static_speed_specouts) as non_tunnel_static_speed_specouts, \
    MAX(n.static_speed_error_duration_max) as static_speed_error_duration_max, \
    MAX(n.static_speed_error_max) as static_speed_error_max, \
    MAX(n.hdg_error_max_exit30s) as hdg_error_max_exit30s, \
    MAX(n.tunnel_out_ffposerr_max) as tunnel_out_ffposerr_max, \
    MAX(n.tunnel_out_ttffs_max) as tunnel_out_ttffs_max \
    FROM gnslogindex g JOIN duterror e ON g.log_id = e.log_id \
    LEFT JOIN navgpssummary n ON n.log_id = g.log_id \
    WHERE  g.competitor \
    {competitor_filter} \
    and g.file_loc in ({file_locations}) \
    and g.date  = '{date}' \
    and e.name in ('_position2DError', 'SOG_Error', '_headingError', '_alongTrackError') \
    group by  g.dut_model, g.dut_num, g.build_num, g.chipset, e.name, g.l5"


    #print(sql)

    df = db.dbFetchAllDF(sql)
    #if DEBUG:
    #    print("-- avg_gnslogindex_duterror_navgps_competitors dataframe generated--")
    #    print(df)
    return df

#------------------------
def add_average_row(df):
#------------------------
    print('add_average_row(): calculate average for the score columns')
    score_cols=[]
    for col in list(df.columns):
       print(col)
       if col.endswith('score'):
         score_cols.append(col)

    #print('score columns are')
    #print(score_cols)
#-----------------------
    # Create an empty list to store the new DataFrame rows
    new_rows = []

    # Group by column 'A'
    for key, group in df.groupby('Test'):
#        print("key=", key) #Heading, etc
#key= 1. Horizontal Error
#key= 2. Tunnel
#key= 3. Speed
#key= 4. VDR
#key= 5. Stop
#key= 6. Heading
        # Append the original rows to new_rows
        new_rows.append(group)

        avg_per_col={}
        for col in score_cols:

            values = group[col].tolist()

            res = [i for i in values if i != '?']

            if len(res) > 0:
               avg_per_col[col] = round (sum(res) / len(res), 2)
            else:
               avg_per_col[col] = None

        # Create a new row with 'Average' in column 'Test' and the calculated averages for score columns
        average_row = {}
        for  col in list(df.columns):
            if col == 'Test':
                average_row[col]=key # + ' average'
            elif col == 'Criteria':
                average_row[col]='Average'
            elif col.endswith('score'):
                average_row[col] = avg_per_col[col]
            else:
                average_row[col]= None

        # Append the average row to new_rows
        new_rows.append(pd.DataFrame([average_row]))

    # Concatenate all the rows (including the new average rows) into a new DataFrame
    new_df = pd.concat(new_rows, ignore_index=True)

    return new_df

#-----------------------
def df_info(name, df):
#-----------------------
    print(name, 'len(df)=', len(df))
    for i, col in enumerate(df.columns):
       print (i, col,  df[col].dtype)

    #print(df)

#---------------------------
def calculate_avg_per_build(df_dut_build, df2, metric):
#---------------------------
 print('calculate_avg_per_build()')
#  Create a mapping of 'dut_num' to 'build_num'
 build_map = df_dut_build.set_index('dut_num')['build_num'].to_dict()
 
 print('build_map=')
 print(build_map)
 score_build_map={}
 for dut, build in build_map.items():
     if not build:
         build = 'REF'
         build_map[dut] = build

     score_build_map[dut+"_score"] = build
     if build == 'REF':
         score_build_map['REF ' + dut+"_score"] = build

 build_map.update(score_build_map)


 if metric == 'Average score': #remove all records which are not Average Score
     df2=df2[df2['Criteria']=='Average']


#  Melt df2 to make it long format for easier processing
 df2_melted = df2.melt(id_vars=['Test','Criteria'], var_name='dut_num', value_name='dut_value')

#  Map the 'dut_num' to the corresponding 'build_num' from the build_map
 if metric == 'Average score':
   df2_melted['build_num'] = df2_melted['dut_num'].map(score_build_map).dropna()
 else:
   df2_melted['build_num'] = df2_melted['dut_num'].map(build_map).fillna('Unknown')


 #remove all records where dun_num not ends with _score

 df2_melted = df2_melted[df2_melted['dut_num'].str.endswith('_score')]
 print('only score df2_melted=')
 #print(df2_melted)


 df2_melted['dut_value'] = pd.to_numeric(df2_melted['dut_value'] , errors = 'coerce')
 #print(df2_melted.dtypes)
# Step 4: Group by 'criteria' and 'build_num', and calculate the mean and count for each group
 df_grouped = df2_melted.dropna(subset=['build_num']).groupby(['Test','Criteria', 'build_num']).agg(
    avg_value=('dut_value', 'mean'), dut_count=('dut_value', 'count')
 ).reset_index()

 print('len df_grouped=', len(df_grouped))
# Test          object
#Criteria      object
#build_num     object
#avg_value    float64
#dut_count      int64

# print(df_grouped.dtypes)
# print(df_grouped)

#                   Test Criteria   build_num   avg_value  dut_count
#0   1. Horizontal Error  Average  4.14.17_7o   83.590000          6
#1   1. Horizontal Error  Average         REF   91.850000          3
#2             2. Tunnel  Average  4.14.17_7o   93.916667          6
#3             2. Tunnel  Average         REF   91.330000          3
#4              3. Speed  Average  4.14.17_7o   81.158333          6
#5              3. Speed  Average         REF   21.113333          3
#6                4. VDR  Average  4.14.17_7o   83.543333          6
#7                4. VDR  Average         REF   32.893333          3
#8               5. Stop  Average  4.14.17_7o  100.000000          6
#9               5. Stop  Average         REF   99.556667          3
#10           6. Heading  Average  4.14.17_7o   98.100000          6
#11           6. Heading  Average         REF   96.876667          3

# Step 5: Pivot the dataframe to have a column for each 'build_num'

 df3 = df_grouped.pivot_table(index=['Test','Criteria'], columns='build_num', values=['avg_value', 'dut_count'], aggfunc={'avg_value':'mean', 'dut_count':'sum'}).reset_index()

 return df3
#-----------------------------------------------
def process_one_location(date, sheet, file_locations, only_builds=None, ignore_builds=None):
#------------------------------------------------
        print('DEBUG process_one_location()')
        print("sheet=", sheet, 'date=', date)

        file_locations_str="'" + file_locations[0] + "'"
        for loc in  file_locations[1:]:
            file_locations_str += ", '" + loc + "'"

        print("file_locations_str=",file_locations_str)
        if POPULATE_KPI_SCORE:
            chipset='ALL'
            competitor_dut_models='ALL'
            non_competitor_dut_models='ALL'
        else:
            chipset=DEFAULT_chipset
            competitor_dut_models=DEFAULT_competitor_dut_models
            non_competitor_dut_models=DEFAULT_non_competitor_dut_models

        if len(file_locations) > 1: # do average (example: 2024-05/2024-05-02/SF_Downtown/Driving/UC/)
            print("DEBUG special case len(file_locations)=", len(file_locations))
            df_competitors     = avg_gnslogindex_duterror_navgps_competitors(date, file_locations_str, competitor_dut_models)
            df_dut     = avg_gnslogindex_duterror_navgps(date, file_locations_str, chipset, non_competitor_dut_models, only_builds, ignore_builds)
        else:
            df_competitors = join_gnslogindex_duterror_navgps_competitors(date, file_locations_str, competitor_dut_models)
            df_dut     = join_gnslogindex_duterror_navgps(date, file_locations_str, chipset, non_competitor_dut_models, only_builds, ignore_builds)

        if df_dut.empty:
                   print(" WARNING df_dut.empty")
        if df_dut.empty and df_competitors.empty:
            print('df_dut.empty and df_competitors.empty')
            return pd.DataFrame(), pd.DataFrame(), pd.DataFrame() #empty dataframe


        print('BEFORE combine df_dut columns =') 

        df = combine_ref_and_dut(sheet, df_dut, df_competitors, date)

        print("AFTER combine_ref_and_dut: df_dut " )

        if POPULATE_KPI_SCORE:
            return None, None, None


        print("--------before drop duplicates-------------df_dut=")
        for col in df_dut.columns:
              print(f"{col}: {df_dut[col].dtype}")
        print("--------------------------")

        print("--------before drop duplicates-------------df_competitors=")
        for col in df_competitors.columns:
              print(f"{col}: {df_competitors[col].dtype}")
        print("--------------------------")

        df_dut_uniq = df_dut[['dut_num','build_num']].drop_duplicates(subset=['dut_num','build_num']).reset_index(drop=True)


        df_ref_uniq = df_competitors[['dut_num','build_num']].drop_duplicates(subset=['dut_num','build_num']).reset_index(drop=True)
        df = add_average_row(df)

        return df, df_dut_uniq, df_ref_uniq
#------------------
def dates_in_range(start, end):
#-----------------
    # Convert the input strings to datetime objects
    start_date = datetime.strptime(start, '%Y-%m-%d')
    end_date = datetime.strptime(end, '%Y-%m-%d')

    # Create an empty list to store the dates
    date_list = []

    # Generate the list of dates between start and end, excluding the end date
    current_date = start_date #+ timedelta(days=1)
    while current_date <= end_date:
        date_list.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)

    return date_list

#--------------------------------------
def fetch_start_end_date_from_build(build):
#--------------------------------------
    sql = f"SELECT min(date) as min_date,  max(date) as max_date FROM gnslogindex WHERE build_num='{build}'"
    #print(sql)
    res = db.dbFetchOne(sql)

    if res and res[0]:
        start_date = str(res[0])
        end_date = str(res[1])
        print('start_date=', start_date, 'end_date=', end_date)
        return start_date, end_date
    else:
        print("Build not found: ", build)
        exit(1)

#--------------------------------
def trending_score_report(args):
#--------------------------------
    n_plots_per_pdf_page = args.plots_per_page
    # Determine the start date
    end_date = None
    if args.folder:
        print('args.folder=', args.folder)
        if 'G2' not in args.folder:
          log("Currently this report process only G2 folder " + str(args.folder))
          exit(1)
    if args.build:
        start_date, end_date = fetch_start_end_date_from_build(args.build)
    elif args.date:
        start_date = str(args.date) #.strftime('%Y-%m-%d')
    elif args.back and not args.date:
        start_date = (datetime.now() - timedelta(days=args.back)).strftime('%Y-%m-%d')
        start_date = str(start_date)
        end_date=str(datetime.now())[:10]
    else:
        log("Error: Either --date or --build must be provided.")
        sys.exit(1)

    # Determine the end date
    if args.end_date:
        end_date = str(args.end_date) #.strftime('%Y-%m-%d')
    elif args.back and args.date: # back from start date
        print('type(args.date)=', type(args.date))
        end_date=str(args.date)
        start_date = (datetime.strptime(end_date, '%Y-%m-%d') - timedelta(days=args.back)).strftime('%Y-%m-%d')
        start_date = str(start_date)
    elif end_date:
        pass
    else: # ???
        log("TODO - unclear case, assign: end_date = start_date ?")
        exit(0)
    if start_date:
        start_date = str(start_date)

    number_of_days = (datetime.strptime(end_date,'%Y-%m-%d') - datetime.strptime(start_date,'%Y-%m-%d')).days
    print("start_date=", start_date, "end_date=", end_date, 'number_of_days=',  number_of_days)
    print(args.only_builds, args.ignore_builds)

    dates = dates_in_range(str(start_date), str(end_date) )
    print(dates)


    df_per_location={}
    for date in dates:
      sheet_to_file_loc = get_sheet_to_file_loc(date)
      if not sheet_to_file_loc or len(sheet_to_file_loc) == 0:
          continue
      for sheet, file_locations in sheet_to_file_loc.items():
          print('sheet=', sheet)
          #if sheet != 'Driving Highway ToSeoul': #DEBUG - one location only
          #    continue
          #if sheet != 'Bundang Urban Driving': #DEBUG - one location only
          #    continue

          df_one_sheet , df_dut, df_ref = process_one_location(date, sheet, file_locations, args.only_builds, args.ignore_builds)
          if POPULATE_KPI_SCORE: # DO NOT GENERATE reports
            continue
 
          if df_one_sheet.empty:
              print('df_one_sheet.empty')
          if df_dut.empty:
              print('df_dut.empty')
          if df_ref.empty:
              print('df_ref.empty')

          df_dut_and_ref = pd.concat([df_dut, df_ref], ignore_index=True, sort = False)
          if df_dut_and_ref.empty:
              print('df_dut_and_ref.empty')

          metric = 'Average score'
          df_avg_per_build = calculate_avg_per_build(df_dut_and_ref, df_one_sheet, metric)

          if df_avg_per_build.empty:
              print('df_avg_per_build.empty')

          df_avg_per_build['date'] = date
          if sheet not in df_per_location:
              df_per_location[sheet] = df_avg_per_build
          else:
              df_per_location[sheet] = pd.concat([df_per_location[sheet], df_avg_per_build], ignore_index=True, sort = False)

    if POPULATE_KPI_SCORE:
        return None, None

    if number_of_days > 30:
        orientation = "Landscape"
    else:
        orientation = "Portrait"

    generated_pdfs=[]
    all_locations = list(df_per_location.keys())

    for location, df_one_location in df_per_location.items():
      if not df_one_location.empty: 
        plot_location(df_one_location, location, IMG_DIR, orientation, n_plots_per_pdf_page)
        output_pdf = OUTPUT_DIR+"\\PDF\\"+location+"_"+start_date+"_"+end_date+".pdf"
        print(output_pdf)
        images_path=IMG_DIR+"/"+location
        make_pdf(images_path, output_pdf, orientation, n_plots_per_pdf_page)
        generated_pdfs.append(output_pdf)
        print("File created", output_pdf)
      else:
        print("WARNING: No data for location", location)


    return generated_pdfs, all_locations
#--------------------------------------------------
def plot_location(df_per_location, location_name, IMG_DIR, orientation, n_plots_per_pdf_page=2, use_10_colors = True ):
#--------------------------------------------------
    # Function to create plots .png files for each location / metric
    print('plot_location() location_name=', location_name, ' df_per_location.empty =', df_per_location.empty)
    if orientation == 'Landscape':
      #width = 8
        height = 8.0 / n_plots_per_pdf_page
        aspect_ratio = 2.68 # 2.75
    else:
        height = 11.0 / n_plots_per_pdf_page # 5.5 # 4
        aspect_ratio = 1.55

    if n_plots_per_pdf_page == 1:
            aspect_ratio = aspect_ratio / 2.0

    width = height  * aspect_ratio

    os.makedirs(IMG_DIR, exist_ok=True)

    # Extract start and end date from the dataframe
    start_date = df_per_location['date'].min()
    end_date = df_per_location['date'].max()
    n_unique_dates = df_per_location['date'].nunique()
    print('n_uniq_dates=', n_unique_dates, 'start_date =', start_date, 'end_date =', end_date)


    builds = [col[1] for col in df_per_location.columns if col[0] == 'avg_value']
    print(builds)
    build_columns = [col for col in df_per_location.columns if col[0] == 'avg_value']
    print(build_columns) # (avg_value, build), ...

    #----------
    # x-axis: build
    #----------

    plt.figure(figsize=(width, height))

    # Test column values: 1. Horizontal Error, 2. Tunnel, 3. Speed, 4. VDR, 5. Stop, 6. Heading
    metric_groups = df_per_location.groupby('Test')
#0 col= ('Criteria', '')   -> it is always the same value: "Average:"
#1 col= ('Test', '')
#2 col= ('avg_value', '4.14.17_7n')
#3 col= ('avg_value', '4.14.17_7o')
#4 col= ('avg_value', 'REF')
#5 col= ('date', '')
#6 col= ('dut_count', '4.14.17_7n')
#7 col= ('dut_count', '4.14.17_7o')
#8 col= ('dut_count', 'REF')

    font_size="12"
# Metric variable below:  1. Horizontal Error, 2. Tunnel, 3. Speed, 4. VDR, 5. Stop, 6. Heading
    for i, (metric, group) in enumerate(metric_groups):

        plt.figure(figsize=(width, height))

        #  Create a new dataframe for aggregation
        results = []

        for build in builds:
        # Calculate weighted average for each build
          avg_value_col = ('avg_value', build)
          dut_count_col = ('dut_count', build)

          # Ensure both columns exist to avoid key errors
          if avg_value_col in group.columns and dut_count_col in group.columns:
            # Weighted average: sum(dut_count * avg_value) / sum(dut_count)
            number_duts_per_build = group[dut_count_col].sum()
            if number_duts_per_build in [0.0, 0]:
                weighted_avg = 0.0
                print(metric, 'Warning: group[dut_count_col].sum() =', number_duts_per_build)
            else:
                weighted_avg = (group[dut_count_col] * group[avg_value_col]).sum() / number_duts_per_build

            results.append({'build': build, 'value': weighted_avg})

        # Convert results to a DataFrame
        df_result = pd.DataFrame(results)

        # Sort by build name, with 'REF' always being the first if it exists
        df_result = df_result.sort_values('build')
        if 'REF' in df_result['build'].values:
           ref_row = df_result[df_result['build'] == 'REF']
           df_result = pd.concat([ref_row, df_result[df_result['build'] != 'REF']])

        plt.plot(df_result['build'], df_result['value'],  color='blue', linewidth=1, marker = 'o', markersize=6, linestyle='-')


        plt.axhline(y=THRESHOLD, color='black', linestyle='dashed')
        plt.title(location_name + " " + metric + " score", fontweight='bold', fontsize=font_size)
        plt.ylabel("Score", fontweight='bold', fontsize=font_size)
        plt.xticks(rotation=45,  fontsize="10", fontweight='bold')
        if (df_result['value'] < -10).any():
            plt.ylim(bottom=-10)
        plt.yticks(fontweight='bold', fontsize=font_size)
        plt.xlabel("")
        plt.grid(True)

    # Save as a PNG file

        plt.tight_layout()
        os.makedirs(IMG_DIR + "/" + location_name, exist_ok=True)
        location_without_spaces=location_name.replace(' ', '_')
        metric_without_spaces = metric.replace(' ', '_') # use _ instead space in file name
        filename = f"{IMG_DIR}/{location_name}/build_trend_{location_without_spaces}_{metric_without_spaces}_{start_date}_{end_date}.png"
        plt.savefig(filename)
        plt.close()
        print(filename)

    #-----------
    # x-axis: date
    #-----------
    #https://matplotlib.org/stable/api/markers_api.html
    markers = ['o', 'P', 's', 'D', '^', 'v', '<', '>', 'x', '+', '*', 'h','X']

    available_10_colors = list(mcolors.TABLEAU_COLORS.values())
    # Remove red from the available colors
    red = mcolors.TABLEAU_COLORS['tab:red']
    available_10_colors.remove(red)

        #unique_names = df['name'].unique()
    non_ref_builds = [build for build in build_columns if build != ('avg_value','REF')]
    tab20 = cm.get_cmap('tab20', len(non_ref_builds))
    non_red_colors = [tab20(k) for k in range(20) if k != 6]
    build_colors = {}
    build_colors[('avg_value','REF')] = 'red'
    for  k, build in enumerate(non_ref_builds):
        print(k, 'build=', build,  non_red_colors[k] )
        build_colors[build] = non_red_colors[k % len(non_red_colors)] # tab20(k)

    for i, (metric, group) in enumerate(metric_groups):
        plt.figure(figsize=(width, height))
        print(metric)

        # Plot each column with appropriate color
        cutoff_negative_y_vals=False
        max_y_value = max(group[column].max() for column in build_columns)
        print('max_y_value =', max_y_value)

        for j, column in enumerate(build_columns):
            if column == ('avg_value','REF'):
              plt.plot(group['date'], group[column], label=column[1], color='red', linewidth=1, marker = 'o', markersize=6, linestyle='-') #, alpha=0.7)
              if (group[column] < -10).any():
                  cutoff_negative_y_vals = True
              #plt.plot(group['date'], group[column], label=column[1], color='red', linewidth=1, marker = 'o',markersize=8, markeredgewidth=2, markeredgecolor='black')
            else:
              marker = markers[j % len(markers)]
              #marker = 'o'
            # Assign non-red colors from the available color list
              if len(build_columns) < 11 or use_10_colors:
                  color = available_10_colors[j % len(available_10_colors)]
              else:
                  color=build_colors[column]

              #y_offset=0.0 #random.random() - 0.5 # to solve overlapping dots issue
              plt.plot(group['date'], group[column], label=column[1], color=color, linewidth=1, marker = marker, markersize=6, linestyle='-') #, alpha=0.7)
              #plt.plot(group['date'], group[column] + y_offset, label=column[1], color=color, linewidth=1, marker = 'o',markersize=8, markeredgewidth=2, markeredgecolor='black')
              if (group[column] < -10).any():
                  cutoff_negative_y_vals = True

        plt.axhline(y=THRESHOLD, color='black', linestyle='dashed')
        plt.title(location_name + " " + metric + " score", fontweight='bold', fontsize=font_size)
        plt.ylabel("Score", fontweight='bold', fontsize=font_size)

        #plt.legend( loc='best', fontsize="10")
        plt.legend( loc='upper left', bbox_to_anchor=(1.05,1), fontsize="8", borderaxespad=0.)
        #plt.xticks(rotation=45, fontweight='bold', fontsize="10")
        plt.xticks(rotation='vertical',  fontsize="8", fontweight='bold')
        plt.yticks(fontweight='bold', fontsize=font_size)
#        plt.autoscale(enable=False, axis='y')
#        if cutoff_negative_y_vals:
#            plt.ylim(bottom=-10)

        plt.xlabel("")
        plt.grid(True)

        if cutoff_negative_y_vals:
            # TODO max_y_value = max(df[column].max() for column in build_columns)
            plt.autoscale(enable=False, axis='y')
            ymin, ymax = plt.gca().get_ylim()
            print('ymin=', ymin, 'ymax=',ymax)
            print('max_y_value =', max_y_value)
            if max_y_value < ymax and ymax > 100:
               plt.ylim(top=100)
            #exit(0)
            plt.ylim(bottom=-10)

    # Save the line plot as a PNG file

        plt.tight_layout()
        os.makedirs(IMG_DIR + "/" + location_name, exist_ok=True)
        location_without_spaces=location_name.replace(' ', '_')
        metric_without_spaces = metric.replace(' ', '_') # use _ instead space in file name
        filename = f"{IMG_DIR}/{location_name}/time_trend_{location_without_spaces}_{metric_without_spaces}_{start_date}_{end_date}.png"
        #line_plot_filename = f"{IMG_DIR}/{location_name}/lineplot_{metric}_{start_date}_{end_date}.png"
        plt.savefig(filename)
        plt.close()
        print("Created file:", filename)

#------------------------------------------

import os
from reportlab.lib.pagesizes import letter, landscape, portrait
from reportlab.pdfgen import canvas
from PIL import Image
from reportlab.lib.units import inch

#--------------------------------------------------------------
def make_pdf(images_folder_path, output_pdf, orientation='Portrait',  n_plots_per_pdf_page=2):
#--------------------------------------------------------------
    # Get all PNG files in the folder and subfolders
    print("make_pdf() images folder_path=", images_folder_path, output_pdf)
    image_files = []
    for root, dirs, files in os.walk(images_folder_path):
        time_trend_files = sorted([images_folder_path + "/"+ f for f in files if f.startswith("time_trend") and f.endswith(".png")])
        build_trend_files =  sorted([images_folder_path + "/"+ f for f in files if f.startswith("build_trend")  and f.endswith(".png")])

    # Combine time_trend files followed by build_trend files
    image_files = time_trend_files + build_trend_files

    # Create a canvas for the PDF
    c = canvas.Canvas(output_pdf, pagesize=landscape(letter))  # Initial page size; will adjust per image

    # Define margins and spacing
    margin = 0.5 * inch  # 1 inch margin
    spacing = 0.5 * inch  # 0.5 inch spacing between images

    # Initialize page number
    page_number = 1

    # Initialize image counter
    img_counter = 0

    for image_path in image_files:
        # Open the image to get its size
        img = Image.open(image_path)
        img_width, img_height = img.size
        aspect_ratio = img_width / img_height

        # print(img_width, img_height,'aspect_ratio=', aspect_ratio) # 600, 400
        # Determine orientation based on aspect ratio
        #if False and aspect_ratio > 1:  # Landscape
        if orientation =='Landscape':
            page_orientation = landscape(letter)
            print('landscape')
        else:  # Portrait
            page_orientation = portrait(letter)
            print('portrait')


        # Set the page size
        c.setPageSize(page_orientation)
        page_width, page_height = page_orientation
        print('page_width=', page_width, ' page_height=',page_height)
        # page_width= 612.0  page_height= 792.0
        # Calculate image size to fit within half the page (two per page)
        #if False and aspect_ratio > 1:  # Landscape image on landscape page
        if orientation =='Landscape':
            img_display_width = (page_width - 2 * margin - spacing) / 2
            img_display_height = img_display_width / aspect_ratio
            img_display_width = 750
            img_display_height = 2 * 280 / n_plots_per_pdf_page
        else:  # Portrait image on portrait page
            #img_display_width = (page_width - 2 * margin - spacing) / 2
            #img_display_height = img_display_width / aspect_ratio
            img_display_width = 550
            img_display_height = 2 * 350 / n_plots_per_pdf_page  # img_display_width / aspect_ratio

        print('img_width=',img_width, 'img_height=', img_height,'aspect_ratio=', aspect_ratio, 'img_aspect_ratio=',img_display_width / img_display_height )
        print('img_display_width=', img_display_width, ' img_display_height=',img_display_height)

        if n_plots_per_pdf_page == 1:
            y_position = margin
        else: # 2 plots per page
          # Alternate between top and bottom positions if 2 plots per page
          position_index = img_counter % 2
          if position_index == 0:
            #y_position = page_height - margin - img_display_height - 0.25 * inch  # Leave space for text
            y_position = page_height /2
          else:
            y_position = margin  # Bottom image

        # Add the image filename above the image
        file_name = os.path.basename(image_path)
        c.setFont("Helvetica-Bold", 12)
        #text_x = margin + (img_display_width + spacing) * (img_counter % 2)  # 672??
        #if aspect_ratio > 1:
            # Center the text above the image
        #    text_x = margin + (img_display_width / 2)
        #y=(page_height - 5.0)/2 * (img_counter % 2)
        #p_x = 336 # text_x + img_display_width / 2
        #page_width, page_height
        p_y = y_position + img_display_height + 0.15 * inch  # 748 or 396
        print('p_y=', p_y , 'p_y-img_display_height-10=', p_y-img_display_height-10)
        #print ('text_x=', text_x, ' file_name=', file_name, 'p_x=',p_x, 'p_y=',p_y)

        #c.drawCentredString(p_x, p_y, file_name)


        # Draw the image
        #print('img_display_width=',img_display_width,' img_display_height=',img_display_height)
        x = margin # + (img_display_width + spacing) * (img_counter % 2)
        #print('img_counter=', img_counter, ' x=',x,' y=', y_position, ' file_name=', file_name)
        c.drawImage(image_path,
                   x,
                   p_y-img_display_height-10,
                   #y_position,
                   width=img_display_width,
                   height=img_display_height)

        img_counter += 1

        # If two images have been added, add footer and move to next page
        if img_counter % 2 == 0 or n_plots_per_pdf_page == 1:
            # Add footer with page number
            c.setFont("Helvetica", 10)
            footer_text = f"Page {page_number}"
            c.drawRightString(page_width - margin, margin / 2, footer_text)

            # Finish the page
            c.showPage()
            page_number += 1

    # If there's an odd number of images, finalize the last page
    if img_counter % 2 != 0 and n_plots_per_pdf_page == 2:
        # Add footer with page number
        c.setFont("Helvetica", 10)
        footer_text = f"Page {page_number}"
        c.drawRightString(page_width - margin, margin / 2, footer_text)

        # Finish the last page
        c.showPage()

    # Save the PDF
    c.save()


#-----------------------------------
def get_sheet_to_file_loc(date):
#-----------------------------------
    # map location -> MS Excel sheet(s)
    known_locations={
    'G2_ToSeoul/Driving' : ['Driving Highway ToSeoul'], #, 'Driving Highway VDR ToSeoul'],
    'G2_ToSuwon/Driving' : ['Driving Highway ToSuwon'], #, 'Driving Highway VDR ToSuwon'],
    'G2_Bundang/Driving' : ['Bundang Urban Driving'],
    #'SF_Downtown/Driving/UC' : ['Driving - SF UC'],
    #'G2_Gangnam'         : ['Pedestrian - Urban'],
    'G2_BundangParkingGarage' : ['Bundang Parking Garage'],
    'G2_SuwonStation'         : ['Suwon Station']
    }

    # FOR DEBUGGING ONLY
    #known_locations={
    #'G2_ToSuwon/Driving' : ['Driving Highway ToSuwon']
    #}

    locations = list(known_locations.keys())

    ref_file_locations = get_ref_file_loc(date, locations, DEFAULT_competitor_dut_models)
    if not ref_file_locations:
        print("There are no competitors data in SPOT-ACE database for", date)


    dut_file_locations = get_dut_file_loc(date, locations, DEFAULT_chipset, DEFAULT_non_competitor_dut_models)
    if not dut_file_locations:
        print("There are no DUTs data in SPOT-ACE database for", date)


    # common_file_locations=list(set(ref_file_locations).intersection(dut_file_locations))
    common_file_locations=list(set(ref_file_locations).union(dut_file_locations))
    if not common_file_locations:
        print("There are no common DUT / REFERENCE data in SPOT-ACE database for", date)
        return {} #empty dict

    print("found locations")
    for c in common_file_locations:
        print(c)

    # for every sheet populate the list of file locations:
    sheet_to_file_loc=defaultdict(list)

    for file_loc in common_file_locations:
        for folder, sheets in known_locations.items():
            if folder in file_loc:
                print(folder, "check if sheets exists", sheets)
                for sheet in sheets:
                    sheet_to_file_loc[sheet].append(file_loc)
                break

    for sheet, sheet_file_locations in sheet_to_file_loc.items():
        print (sheet, "locations=", sheet_file_locations)

    print("Number of spreadsheets=", len(sheet_to_file_loc))
    return sheet_to_file_loc


#-----------------------------------
def daily_score_report(date,  output_excel_file):
#-----------------------------------
# TODO: inside get_sheet_to_file_loc()   we are using DEFAULT_chipset - how to adjust it for all chipsets???
    sheet_to_file_loc = get_sheet_to_file_loc(date) 
    if len(sheet_to_file_loc) == 0:
        return None, None

    all_dfs_per_location={}
    all_dut_per_location={}
    if not  POPULATE_KPI_SCORE:
      wb = Workbook()
      ws = wb.active

    for sheet, file_locations in sheet_to_file_loc.items():
        print("sheet", sheet)
        df_one_sheet, df_dut_uniq, df_ref_uniq = process_one_location(date, sheet, file_locations)
        if POPULATE_KPI_SCORE: # DO not create reports
             continue

        all_dfs_per_location[sheet] = df_one_sheet
        all_dut_per_location[sheet] = df_dut_uniq
        ws = wb.create_sheet(title=sheet)
        populate_sheet(ws, df_one_sheet, date, df_dut_uniq)

    if not  POPULATE_KPI_SCORE:
        del wb['Sheet']
        wb.save(output_excel_file)
        wb.close()

    return all_dfs_per_location, all_dut_per_location


def valid_date(date_str):
    """Validate the date format is YYYY-MM-DD."""
    #log("Validate input date: "+ date_str)
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        log("Error in input date: "+ date_str)
        raise argparse.ArgumentTypeError(f"Invalid date format: '{date_str}'. Expected format: YYYY-MM-DD.")

def parse_arguments():
    parser = argparse.ArgumentParser(description="Fetch records from PostgreSQL table T.")

    # Adding command line arguments
    parser.add_argument('--date', type=valid_date, help='Start date in format YYYY-MM-DD.')
    parser.add_argument('--end_date', type=valid_date, help='End date in format YYYY-MM-DD.')
    parser.add_argument('--back', type=int, help='Number of days back from --date.')
    parser.add_argument('--build', type=str, help='Build name to detect the start date.')
    parser.add_argument('--folder', type=str, help='Input folder.')
    parser.add_argument('--ignore_builds', nargs='*', help='List of builds to ignore.')
    parser.add_argument('--only_builds', nargs='*', help='List of builds to process.')
    parser.add_argument('-o', '--output_dir',  required=False, default=OUTPUT_DIR, help = " default: %(default)s")
    parser.add_argument('--send_email', action='store_true', help='Send e-mail if this flag is provided')
    parser.add_argument('--plots_per_page', type=int, choices=[1,2], default=1, help='Number of plots per PDF page (default=1, other option: 2')
    parser.add_argument('--kpi_score',  action='store_true', help="Populate kpi_score table")
    parser.add_argument('--kpi_score_gdc',  action='store_true', help="Populate kpi_score_gdc and kpi_score_gdc_summary tables")
    parser.add_argument('--kpi_g2_pdf',  action='store_true', help="Create MS Excel and PDFs with kpi trends for G2")
    args = parser.parse_args()

    print(args.kpi_score)
    print(args.kpi_score_gdc)
    print(args.kpi_g2_pdf)

    mutually_exclusive_options=args.kpi_score + args.kpi_score_gdc + args.kpi_g2_pdf
    if mutually_exclusive_options !=1 :
       print("One of the following options need to be provided: --kpi_score --kpi_score_gdc --kpi_g2_pdf")
       sys.exit(0)

    # Validations
    if args.date and args.build:
        log("Error: --date and --build cannot be provided together.")
        sys.exit(1)

    if args.back and args.end_date:
        log("Error: --back and --end_date cannot be provided together.")
        sys.exit(1)

    return args
#-----------------------------------
def send_mail(list_of_attachments, html_body):
#-----------------------------------
    print(" send_mail() start")
    RECIPIENTS_DEVELOPER_MODE = 'jiwei.wei@samsung.com, m.lubinsky@partner.samsung.com, kanepan@samsung.com, jl.deberg@samsung.com, sundar.raman@samsung.com, bhaskar.n@samsung.com, ranjan.rao@samsung.com, bholterh@samsung.com, ksoo.yu@samsung.com, jungbeom.kim@samsung.com, andrew2.lee@samsung.com'
    #RECIPIENTS_DEVELOPER_MODE = 'jiwei.wei@samsung.com, kanepan@samsung.com, m.lubinsky@partner.samsung.com'
    #RECIPIENTS_DEVELOPER_MODE = 'm.lubinsky@partner.samsung.com'

    print('list_of_attachments=')
    print(list_of_attachments)

    if list_of_attachments:
        attachments = ','.join(list_of_attachments)
        body="MX score trend reports are attached."
    else:
        attachments = ''
        body="No attachments."

    print('body=', body)

    email_engine=SpotEmailNew(sender=None, recipients=RECIPIENTS_DEVELOPER_MODE)
    print(" send_mail() before")
    email_engine.send_email(subject="MX Score Trend Report",
                            text=body,
                            files=attachments,
                            html=html_body
                            )
    print(" send_mail() end")
#--------------------------------
def generate_fail_total_table(metric_score):
#--------------------------------
    table_html = '<table border="1" style="border-collapse: collapse;">'
    table_html += '<tr><th>Score Metric</th><th>Status</th><th>DUT Fails/Total</th></tr>'
    for k, v in metric_score.items():
        status=v[0]
        fails_total = v[1]
        table_html += f'<tr><td>{k}</td><td>{status}</td><td>{fails_total}</td></tr>'
    table_html += '</table>'
    return table_html


#--------------------------------
def generate_dut_table(dut_build):
#--------------------------------
    table_html = '<table border="1" style="border-collapse: collapse;">'
    table_html += '<tr><th>DUT</th><th>Build</th></tr>'
    for dut, build in dut_build.items():
        table_html += f'<tr><td>{dut}</td><td>{build}</td></tr>'
    table_html += '</table>'
    return table_html

#-----------------------
def get_dut_scores(df):
#-----------------------
# extract rows where 'Criteria' = 'Average' and ignore columns which starts from REF, keep only columns which ends with score
    """
return dictionary like this:
{  "horizontal": ("Fail",  "1/6"),
   "heading":   ("Pass",  "0/6")
}
    """
    df_average = df[df['Criteria'] =='Average']
    columns_of_interest = [col for col in df.columns if col.endswith('score') and not col.startswith('REF')]
    n_dut_total = len(columns_of_interest)
    if n_dut_total == 0:
         return None

    result={}
    for index, row in df_average.iterrows():
        # Extract the unique value in the 'Test' column
        test_value = row['Test']

        # The number of columns of interest where the value is below the threshold T
        n_dut_failed = sum(row[col] < THRESHOLD for col in columns_of_interest)

        # Determine the status: 'Pass' if N_1 == 0, else 'Fail'
        status = 'Pass' if n_dut_failed == 0 else 'Fail'

        ratio = f"{n_dut_failed} / {n_dut_total}"

        # Assign the result in the dictionary
        result[test_value] = (status, ratio)


    #print(result)
    return result


#-----------------
def create_html(locations, all_dfs_per_location, all_dut_per_location, start_date):
#-----------------
    location_scores={}
    location_images={}

    for location in locations:
        if location not in all_dfs_per_location:
            print("ERROR: no dataframe for ", location)
            exit(0)
        else:
           location_scores[location] = get_dut_scores(all_dfs_per_location[location])


        #location_scores[location] = {"horizontal": ("Fail",  "1/6"),
        #                              "heading":   ("Pass",  "0/6") }

        img_folder = os.path.join(IMG_DIR,location)
        # TODO instead code below think is it possible to re-use get_images_for_email()
        for root, dirs, files in os.walk(img_folder):

            for f in files:

                if "Horizontal" in f:
                  if location not in location_images:
                      location_images[location]=[]
                  location_images[location].append(os.path.basename(f))

            ## Sort images: time_trend before build_trend
            if location in location_images:
                location_images[location].sort(reverse=True)


    body_html = '<html><body>'

    for location, metric_score in location_scores.items():

        print('location=', location)
        print('metric_score=', metric_score)
        if metric_score is None:
            continue
        body_html += f'<h3>{location}  {start_date}</h3>'

        # Add the table for metric_score
        body_html += generate_fail_total_table(metric_score)

        body_html += '<br><br>'
        # Add the table for dut build
        dut_build_dict = all_dut_per_location[location].set_index('dut_num')['build_num'].to_dict()
        sorted_dict = dict(sorted(dut_build_dict.items()))
        body_html += generate_dut_table(sorted_dict)

        # images
        if location in location_images:
            for fname in location_images[location]:
                body_html += f'<p><img src="cid:{fname}" alt="{fname}" style="max-width:500px;"></p>'

    body_html += '</body></html>'
    print(body_html)
    return body_html
#-----------------------------
def get_images_for_email(all_locations, pattern="Horizontal"):
#--------------------------
    images=[]
    for location in all_locations:
        img_folder = os.path.join(IMG_DIR,location)
        print(img_folder)
        for root, dirs, files in os.walk(img_folder):
            for f in files:
                print(f)
                if pattern in f:
                    full_name = os.path.join(img_folder,f)
                    print(full_name)
                    images.append(full_name)
    return images


#--------------------------
def upload_errorDF(date):
#--------------------------
    date=str(date)
    print('upload_errorDF() date=', date, type(date), str(date))
    dir=r"V:\RegressionLibrary\SPOT-ACE"
    dir=os.path.join(dir, date[0:7])
    print(dir)
    dir=os.path.join(dir, date)
    print(dir)
    if not os.path.exists(dir):
        print("not exists", dir)
        return

    total_lines = 0
    files_count = 0
    # Walk through all subdirectories of the given directory
    for root, dirs, files in os.walk(dir):
        # Check if the current folder is named 'lsi'
        if os.path.basename(root) == "lsi":
            for file in files:
                if file.endswith("errorDF.csv"):
                    file_path = os.path.join(root, file)
                    files_count += 1
                    try:
                        with open(file_path, 'r', encoding='utf-8') as f:
                            line_count = sum(1 for _ in f)
                            total_lines += line_count
                    except Exception as e:
                        print(f"Error reading {file_path}: {e}")

    print(f"Total lines: {total_lines}  Total files: {files_count}")

#-----------------------------------
def find_gdc_score_files(folder):
#-----------------------------------
    gdc_score_files = {}
    
    for root, dirs, files in os.walk(folder):
        #if ( "GDC_DEV" in root and 'G2' in folder ) or ("GDC_Output" in root and 'SF' in folder):
        #if "GDC_Output" in root: # and 'LowSpeed' in root:
        #if "GDC_Output" in root and ("DUT_ALL" not in root) and ("SeoulTunnel" in root): # and ( "Diablo_Driving" in root) :
        if "GDC_Results" in root:
            score_files = [f for f in files if f.endswith("_score.csv")]
            for score_file in score_files:
                score_path = os.path.join(root, score_file)
                print('score_paths=',score_path)
                for_mx_dashboard = None

                #two_levels_up = os.path.dirname(os.path.dirname(root))
                level_up = os.path.dirname(root)
                for sub_root, sub_dirs, sub_files in os.walk(level_up):
                    if "for_mx_dashboard.csv" in sub_files:
                        for_mx_dashboard = os.path.join(sub_root, "for_mx_dashboard.csv")
                        print('for_mx_dashboard=', for_mx_dashboard)
                        break
                gdc_score_files[score_path] = for_mx_dashboard
    
    return gdc_score_files

#----------------------------
def get_last_line_number_in_top_section(fname):
#----------------------------
    with open(fname, 'r', encoding='utf-8') as f:
       for i, line in enumerate(f, start=0):
            if (i > 1) and ('TEST' in line) and ('Category' in line) and ('ITEM' in line):
              return i , line
    print("Error - cannot find the last line in top section ", fname)    
    return None

#-------------------------------------------------
def read_top_section(score_file, last_line_number_in_top_section):
#-------------------------------------------------
    df = pd.read_csv(score_file, nrows=last_line_number_in_top_section - 1, na_values=' ')
    
    df.drop('Unnamed: 6', axis=1, inplace=True)
    df.columns = [ 'VAL_SCORE_AVG' if (not col.strip() and i==3) else col
                  for i, col in enumerate(df.columns)]

    df.drop('TEST', axis=1, inplace=True)

    df.loc[df["VAL_SCORE_AVG"] == "AVG", "ITEM"] = "_Average"
    df.loc[(df["VAL_SCORE_AVG"] == "AVG") & (df["VAL_SCORE_AVG"].shift()== "AVG"), "ITEM"] = "__Summary"
    # We use double undescore __ in order to ORDER BY criteria DESC sort to work: _Average before __Summary
    # TODO: for the _Summary change the Category  to be previous row Category (N + 1). SUMMARY

    df[['Category', 'ITEM']] = df[['Category', 'ITEM']].fillna(method='ffill')

    df_avg = df[df["VAL_SCORE_AVG"] == "AVG"].copy()

    df_value_score = df[df["VAL_SCORE_AVG"].isin(["Value", "Score"])].pivot_table(
        index=["Category", "ITEM"],
        columns="VAL_SCORE_AVG",
        values=["REF", "DUT"],
        aggfunc="first" 
    )
 
    df_value_score.columns = [f"{col[0]}_{col[1]}" for col in df_value_score.columns]
    df_value_score.reset_index(inplace=True)

    df_avg.rename(columns={"REF": "REF_Score", "DUT": "DUT_Score"}, inplace=True)
    df_avg.drop(columns=["VAL_SCORE_AVG"], inplace=True)
    
    # Concat both dataframes

    df_top = pd.concat([df_avg, df_value_score], ignore_index=True)

    cols=['REF_Score', 'DUT_Score', 'DUT_Value', 'REF_Value']
    

    print(df_top.info())
    if df_top.empty:
        print( 'df_top empty')
        exit(0)
    missing_cols = [col for col in cols if col not in df_top.columns]
    if missing_cols:
        print("missing cols", missing_cols)
        exit(0)



    #df_top[cols] = df_top[cols].replace('-', np.nan)  # another option: remove such records
    dtypes = { col: df_top[col].dtype for col in cols }
    if all (dtype == 'object' for dtype in dtypes.values()):
        if df_top[cols].isna().any().any():
          print("Found None values")
          #print(df_top[cols])
          for col in cols:
            #print('col=', col)
            df_top[ col ] = df_top[col].fillna('')

        mask=pd.Series(True, index = df_top.index)
        for i, col in enumerate(cols):
            print(i, col, "build mask")
            mask &= ~df_top[col].str.contains('-', na=False)

        df_top = df_top[mask]    
        print(df_top[cols])
        df_top[cols] = df_top[cols].replace('',np.nan)
        df_top[cols] = df_top[cols].astype(float)

    df_top.sort_values(by=["Category","ITEM"], ascending=[True, True], inplace=True)

    df_top = df_top.rename(columns={
                'Category': 'test', 
                'ITEM'      : 'criteria',
                'REF_Value' : 'ref_val', 
                'REF_Score' : 'ref_score', 
                'DUT_Value' : 'dut_val', 
                'DUT_Score' : 'dut_score'
            })

    #   if ref_score or dut_score = 0 then replace it with 100
    print('df_top.info()=')
    print(df_top.info())
    print(df_top)
    print(df_top['ref_score'])
    df_top.loc[df_top['ref_score'] < 0.001,  'ref_score'] = 100.00   ## there are -1 values as well
    df_top.loc[df_top['dut_score'] < 0.001,  'dut_score'] = 100.00
    #column_order = ['date', 'test_loc', 'test_loc2', 'test_type', 'output_type', 'test', 'criteria', 'ref_val', 'ref_score', 'dut_val', 'dut_score']
    #df_top_final = df_top[column_order]

    return(df_top)

#-------------------------------------
def read_bottom_section(fname, line):
#-------------------------------------
  print('read_bottom_section line=', line )
  df = pd.read_csv(fname, skiprows=line, na_values=' ' )
  print(df.info())
  print(df)

  df.columns = [ 'VAL_SCORE_AVG' if (not col.strip() and i==3) else col
                  for i, col in enumerate(df.columns)]

  df.drop('TEST', axis=1, inplace=True)
  columns_to_drop = [col for col in df.columns if col.startswith("Unnamed")]
  df = df.drop(columns=columns_to_drop)

  df.loc[df["VAL_SCORE_AVG"] == "AVG", "ITEM"] = "_Average"
  df.loc[(df["VAL_SCORE_AVG"] == "AVG") & (df["VAL_SCORE_AVG"].shift()== "AVG"), "ITEM"] = "__Summary"
  # TODO we do not need to import this row "_Summary" to database: remove this reodr from dataframe
  df[['Category', 'ITEM']] = df[['Category', 'ITEM']].fillna(method='ffill')


  df_avg = df[df["VAL_SCORE_AVG"] == "AVG"].copy()


  ref_dut_columns = [col for col in df.columns if col.startswith(('(REF)', '(DUT)'))]

  print("\n\n STEP 1")
  df_melted = df.melt(id_vars=['Category', 'ITEM', 'VAL_SCORE_AVG'], value_vars=ref_dut_columns, var_name='REF_DUT_col', value_name='Value')

   # Extract the suffix from column names ((REF)* or (DUT)*)
  df_melted['dut'] = df_melted['REF_DUT_col'].str[5:]  # Extract after '(REF)' or '(DUT)'

  #print("DEBUG B6_F741U#01_backseat_SM-F741U1-TRACKING")
  #df_melted = df_melted[df_melted["dut"] == "B6_F741U#01_backseat_SM-F741U1-TRACKING"].copy()

  # Determine if the first letters (REF)
  df_melted['is_ref'] = df_melted['REF_DUT_col'].str.startswith('(REF)')

  print("\n\n STEP 2")
  # Pivot to get 'score', 'val', and 'avg' values in separate columns
  df_pivot = df_melted.pivot_table(index=[  'Category', 'ITEM', 'is_ref', 'dut'],
                                     columns='VAL_SCORE_AVG', values='Value', aggfunc='first').reset_index()


  print("\n\n STEP 3")
  # Rename TYPE columns properly
  df_pivot.columns.name = None  # Remove multi-index
  df_pivot = df_pivot.rename(columns={
    'Score': 'score', 
    'Value': 'val', 
    'AVG'  : 'avg',
    'Category': 'test', 
    'ITEM'    : 'criteria'
    }
    )

  print('!!!  before drop df_pivot=')
  print(df_pivot.info())


  if 'avg' in df_pivot.columns:
    df_avg = df_pivot.dropna(subset=['avg']).copy()  # another way: keep only records where df_pivot["criteria"=='_Average']   or '_Summary'
    df_avg["score"] = df_avg["avg"]
    df_avg["val"] = np.nan
    print('after dropna len(df_avg)=', len(df_avg))

    df_avg.drop('avg', axis=1, inplace=True)
    df_pivot.drop('avg', axis=1, inplace=True)

    #remove from df_pivot _Average and _Summary
    print("Remove _Average")
    df_pivot = df_pivot[df_pivot["criteria"] != "_Average"].copy()
    df_pivot = df_pivot[df_pivot["criteria"] != "__Summary"].copy()
    print('!!!  after drop df_pivot=')
    print(df_pivot)
    print(df_pivot.info())

    df_final = pd.concat([df_avg, df_pivot], ignore_index=True)
  else:
    df_final = df_pivot

  print('\n\n df_final=')
  print(df_final)
  #exit(0)


  ## ORDER BY Category, ITEM, dut
  ## TODO : there are 2 AVG  rows just in 4.Heading section just before TEST INFO | Static Heading (line 69)
  ##--------------------------------------------------------------------------------------------------------

  cols=['score', 'val'] #, 'avg']
  df_final[cols] = df_final[cols].replace('-', np.nan) # another option: remove such rows

  df_final.sort_values(by=["test", "dut", "criteria"], ascending=[True, True, True], inplace=True)
  print('\n \n STEP 4 df_final')
  print(df_final.info()) # columns: Category, ITEM, is_ref, dut, avg, score, val
  print(df_final)
  return df_final


#-------------------------------------------------
def read_top_section_with_group(score_file, last_line_number_in_top_section):
#-------------------------------------------------
    print("read_top_section_with_group() \n score_file=" , score_file)
    df = pd.read_csv(score_file, nrows=last_line_number_in_top_section - 1, na_values=' ')
    for i, col in enumerate(df.columns):
        print(i, col)
        if col.startswith('Unnamed'):
            df.drop(col, axis=1, inplace=True)

    df.columns = [ 'VAL_SCORE_AVG' if (not col.strip() and i==3) else col
                  for i, col in enumerate(df.columns)]

    df.drop('TEST', axis=1, inplace=True)

    df.loc[df["VAL_SCORE_AVG"] == "AVG", "ITEM"] = "_Average"
    df.loc[(df["VAL_SCORE_AVG"] == "AVG") & (df["VAL_SCORE_AVG"].shift()== "AVG"), "ITEM"] = "__Summary"
    # We use double undescore __ in order to ORDER BY criteria DESC sort to work: _Average before __Summary
    # TODO: for the _Summary change the Category  to be previous row Category (N + 1). SUMMARY

    df[['Category', 'ITEM']] = df[['Category', 'ITEM']].fillna(method='ffill')

    df_avg = df[df["VAL_SCORE_AVG"] == "AVG"].copy()

    ref_dut_columns = [col for col in df.columns if col not in ['TEST','Category','ITEM','VAL_SCORE_AVG']]
    print('ref_dut_columns=')
    print(ref_dut_columns)
    print("\n\n STEP 1")
    df_melted = df.melt(id_vars=['Category', 'ITEM', 'VAL_SCORE_AVG'], value_vars=ref_dut_columns, var_name='group_name', value_name='Value')
    print(df_melted.info())
    print(df_melted)
    # Determine if the first letters (REF)
    df_melted['is_ref'] = df_melted['group_name'].str.startswith('REF')

    print("\n\n STEP 2")
    # Pivot to get 'score', 'val', and 'avg' values in separate columns
    df_pivot = df_melted.pivot_table(index=['Category', 'ITEM', 'is_ref', 'group_name'],
                                     columns='VAL_SCORE_AVG', values='Value', aggfunc='first').reset_index()

    print(df_pivot.info())
    print(df_pivot)

    print("\n\n STEP 3")
    # Rename  columns properly
    df_pivot.columns.name = None  # Remove multi-index
    df_pivot = df_pivot.rename(columns={
    'Score': 'score', 
    'Value': 'val', 
    'AVG'  : 'avg',
    'Category': 'test', 
    'ITEM'    : 'criteria'
    }
    )

    print('!!!  before drop df_pivot=')
    print(df_pivot.info())

    print("\n\n STEP 4")
    if 'avg' in df_pivot.columns:
      df_avg = df_pivot.dropna(subset=['avg']).copy()  # another way: keep only records where df_pivot["criteria"=='_Average']   or '__Summary'
      df_avg["score"] = df_avg["avg"]
      df_avg["val"] = np.nan
      print('after dropna len(df_avg)=', len(df_avg))

      df_avg.drop('avg', axis=1, inplace=True)
      df_pivot.drop('avg', axis=1, inplace=True)

      #remove from df_pivot _Average and __Summary
      print("Remove _Average")
      df_pivot = df_pivot[df_pivot["criteria"] != "_Average"].copy()
      df_pivot = df_pivot[df_pivot["criteria"] != "__Summary"].copy()
      print('!!!  after drop df_pivot=')
      print(df_pivot)
      print(df_pivot.info())

      df_final = pd.concat([df_avg, df_pivot], ignore_index=True)
    else:
      df_final = df_pivot

    print('\n\n df_final=')
    print(df_final)

    cols=['score', 'val'] #, 'avg']
    column_type = df_final['score'].dtype
    print('score column_type', column_type)
    if column_type == 'object':
      df_final = df_final[~df_final['score'].str.contains('-', na=False)]
      df_final['score'] = df_final['score'].astype(float)
    #df_final = df_final[df_final['val'] != '-']
    column_type = df_final['val'].dtype
    print('val column_type', column_type)
    if column_type == 'object':
      df_final = df_final[~df_final['val'].str.contains('-', na=False)]
      df_final['val'] = df_final['val'].astype(float)
    #df_final[cols] = df_final[cols].replace('-', np.nan) # happens in [VDR]. Another option: remove such rows from df


    print ("before applying filter")
    print(df_final)


    df_final.loc[ (df_final['score'] < 0.001) & (df_final['score'] > -0.001),  'score'] = 100.00 # WARNING can be -1


  ##There are 2 AVG  rows just in 4.Heading section just before TEST INFO | Static Heading 
  ##--------------------------------------------------------------------------------------------------------



    df_final.sort_values(by=["test", "group_name", "criteria"], ascending=[True, True, True], inplace=True)
    print('\n \n STEP 4 df_final len(df_final)=', len(df_final))
    print(df_final.info()) # columns: Category, ITEM, is_ref, dut, avg, score, val
    print(df_final)

    #print("EXIT FOR NOW")
    #exit(0)
    return df_final

#------------------------------------------------
def extract_substring(row, start_char, end_char):
#------------------------------------------------
    text = row['REF_DUT_col']
    start_index = text.find(start_char)

    if start_index == -1:
        return pd.NA
    if not end_char:
       return text[start_index+1:]

    end_index = text.find(end_char, start_index + len(start_char))

    if end_index == -1:
        return pd.NA

    return text[start_index + len(start_char):end_index]

#-------------------------------------
def read_bottom_section_with_group(fname, line):
#-------------------------------------
  print('read_bottom_section line=', line )
  df = pd.read_csv(fname, skiprows=line, na_values=' ' )
  print(df.info())
  print(df)
  print('len(df)=', len(df))

  df.columns = [ 'VAL_SCORE_AVG' if (not col.strip() and i==3) else col
                  for i, col in enumerate(df.columns)]

  df.drop('TEST', axis=1, inplace=True)
  columns_to_drop = [col for col in df.columns if col.startswith("Unnamed")]
  df = df.drop(columns=columns_to_drop)

  df.loc[df["VAL_SCORE_AVG"] == "AVG", "ITEM"] = "_Average"
  df.loc[(df["VAL_SCORE_AVG"] == "AVG") & (df["VAL_SCORE_AVG"].shift()== "AVG"), "ITEM"] = "__Summary"
  # TODO we do not need to import this row "_Summary" to database: remove this 
  df[['Category', 'ITEM']] = df[['Category', 'ITEM']].fillna(method='ffill')

  df_avg = df[df["VAL_SCORE_AVG"] == "AVG"].copy()
  ref_dut_columns = [col for col in df.columns if col.startswith('(')]

  print("\n\n STEP 1 ref_dut_columns=")
  print(ref_dut_columns)
  df_melted = df.melt(id_vars=['Category', 'ITEM', 'VAL_SCORE_AVG'], value_vars=ref_dut_columns, var_name='REF_DUT_col', value_name='Value')
  print(df_melted.info())

   # Extract the group and dut  column name 

  df_melted['group'] = df_melted.apply(lambda row: extract_substring(row, '(', ')'), axis=1)
  df_melted['dut'] = df_melted.apply(lambda row: extract_substring(row, ')', None), axis=1)
  df_melted['is_ref'] = df_melted['REF_DUT_col'].str.startswith('(REF')
  print(df_melted.info())
  
  print(df_melted)
  print('len(df_melted)=', len(df_melted))
 
  print("\n\n STEP 2")
  # Pivot to get 'score', 'val', and 'avg' values in separate columns
  df_pivot = df_melted.pivot_table(index=[  'Category', 'ITEM', 'is_ref', 'dut','group'],
                                     columns='VAL_SCORE_AVG', values='Value', aggfunc='first').reset_index()


  print("\n\n STEP 3")
  # Rename  columns properly
  df_pivot.columns.name = None  # Remove multi-index
  df_pivot = df_pivot.rename(columns={
    'Score': 'score', 
    'Value': 'val', 
    'AVG'  : 'avg',
    'Category': 'test', 
    'ITEM'    : 'criteria'
    }
    )

  print('!!!  before drop df_pivot=')
  print(df_pivot.info())
  print('len(df_pivot)=', len(df_pivot))
  #print()

  if 'avg' in df_pivot.columns:
    df_avg = df_pivot.dropna(subset=['avg']).copy()  # another way: keep only records where df_pivot["criteria"=='_Average']   or '_Summary'
    df_avg["score"] = df_avg["avg"]
    df_avg["val"] = np.nan
    print('after dropna len(df_avg)=', len(df_avg))

    df_avg.drop('avg', axis=1, inplace=True)
    df_pivot.drop('avg', axis=1, inplace=True)

    #remove from df_pivot _Average and _Summary
    print("Remove _Average")
    df_pivot = df_pivot[df_pivot["criteria"] != "_Average"].copy()
    df_pivot = df_pivot[df_pivot["criteria"] != "__Summary"].copy()
    print('!!!  after drop df_pivot=')
    print(df_pivot)
    print(df_pivot.info())

    print('len(df_avg)=', len(df_avg))
    print('len(df_pivot)=', len(df_pivot))

    df_final = pd.concat([df_avg, df_pivot], ignore_index=True)
  else:
    df_final = df_pivot

  print('\n\n len(df_final)=', len(df_final))
  print(df_final)
   


  ## ORDER BY Category, ITEM, dut
  ## TODO : there are 2 AVG  rows just in 4.Heading section just before TEST INFO | Static Heading (line 69)
  ##--------------------------------------------------------------------------------------------------------

  
  #df_final = df_final[df_final['score'] != '-']
  #df_final = df_final[df_final['val'] != '-']

  column_type = df_final['score'].dtype
  print('score column_type', column_type)
  if column_type == 'object':
      df_final = df_final[~df_final['score'].str.contains('-', na=False)]
      df_final['score'] = df_final['score'].astype(float)
  column_type = df_final['val'].dtype
  print('val column_type', column_type)
  if column_type == 'object':
      df_final = df_final[~df_final['val'].str.contains('-', na=False)]
      df_final['val'] = df_final['val'].astype(float)
  print('\n\n after removing - len(df_final)=', len(df_final))
  #cols=['score', 'val'] #, 'avg']
  #df_final[cols] = df_final[cols].replace('-', np.nan) # another option: remove such rows

  df_final.sort_values(by=["test", "dut", "criteria"], ascending=[True, True, True], inplace=True)
  print('\n \n STEP 4 df_final')
  print(df_final.info()) # columns: Category, ITEM, is_ref, dut, avg, score, val
  print(df_final)
  return df_final




#-----------------------------------
def left_join_startswith(df1, df2): # df_bottom, df_for_mx_dashboard
#-----------------------------------
    # add columns: chipset, build, model
    result = df1.copy()
    cannot_find_matching_dut=0
    # For each row in df1, find matching rows in df2
    for col in df2.columns:
        #if col != 'dut_num':
        if col  not in ['dut_num', 'dut_model', 'chipset', 'build_num']:
            result[col] = None
    
    # Iterate through and match
    for i, row in df1.iterrows():
        matches = df2[df2['dut_num'].apply(lambda x: row['dut'].startswith(x))]
        if not matches.empty:
            for col in matches.columns:
                if col != 'dut_num':
                    result.loc[i, col] = matches.iloc[0][col]

                result.loc[i, 'dut_num'] = matches.iloc[0]['dut_num']
                if matches.iloc[0]['build_num']:
                    result.loc[i, 'build_num'] = matches.iloc[0]['build_num']
                else:
                    result.loc[i, 'build_num'] = 'cannot find'
                    cannot_find_matching_dut=1

                result.loc[i, 'chipset'] = matches.iloc[0]['chipset']
                result.loc[i, 'dut_model'] = matches.iloc[0]['dut_model']
        else:
                cannot_find_matching_dut=2
                result.loc[i, 'dut_num'] = result.loc[i, 'dut']
                result.loc[i, 'build_num'] = 'cannot find'
                result.loc[i, 'chipset'] =  'cannot find'
                result.loc[i, 'dut_model'] = 'cannot find'
                
    print("afrer left_join_startswith()")
    print(result)
    print(result.info())
    #if  cannot_find_matching_dut > 0:
    #  print("DEBUG exit in left_join_startswith() cannot_find_matching_dut=", cannot_find_matching_dut)
    #  exit(0)
    return result

#--------------------------------------------------------------
def find_non_matching_duts(score_duts, for_mx_dashboard_duts, score_file, for_mx_dashboard_file):
#--------------------------------------------------------------
    print("started find_non_matching_duts()")
    number_of_issues=0
    stop_flag = False
    for score_dut in score_duts:
      if score_dut in ['P25EVT1' , 'P25PROTO' , 'P25PROTO1']:
         stop_flag = True  

      found_count=0
      for for_mx_dashboard_dut  in for_mx_dashboard_duts:
        if score_dut.startswith(for_mx_dashboard_dut):
            found_count +=1

      if found_count == 0:
        log("Not found matching DUT for " + score_dut + " in " + for_mx_dashboard_file)
        number_of_issues +=1
      elif  found_count >  1: 
        log(str(found_count) + "  matching DUT for " + score_dut + " in " + for_mx_dashboard_file)
        number_of_issues +=1  

    if number_of_issues > 0:
        log("There are " + str(number_of_issues) + " DUT matching issues in "+ score_file)

#------------------------------------------------------------
def find_build_per_group(df_for_mx_dashboard, bottom_header):
#------------------------------------------------------------
    print(bottom_header)
    words = bottom_header.split(',')
    group2dut = defaultdict(list)
    for w in words:
        start=w.find("(") + 1
        end = w.find(")")
        if start > 0 and end  > start:
            group = w[start:end]
            dut = w[end+1:]
            print(w, group, dut)
            group2dut[group].append(dut)

    for k, v in group2dut.items():
        print(k,'=',v)


    dut2build = dict(zip(df_for_mx_dashboard['dut_num'] , df_for_mx_dashboard['build_num']))
    for k, v in dut2build.items():
        print(k,'====',v)


    group2build = {}

    for group, duts in group2dut.items():
      matched_builds = []

      for dut in duts:
        for prefix in dut2build:
            if dut.startswith(prefix):
                matched_builds.append(dut2build[prefix])
                break  # stop at the first match

      # Get most common build (if any match found)
      if matched_builds:
        most_common_build = Counter(matched_builds).most_common(1)[0][0]
        group2build[group] = most_common_build

    print(group2build)

    return group2build
#-----------------------------------------------
def process_gdc_kpi(date, score_file, for_mx_dashboard_file):
#-----------------------------------------------
    print("process_gdc_kpi() \n score_file=", score_file)
    print("process_gdc_kpi() \n for_mx_dashboard_file=", for_mx_dashboard_file)

    if not for_mx_dashboard_file:
        print("ERROR - cannot find for_mx_dashboard.csv file for ", score_file)
        #exit(0) # DEBUG
        return

    full_folder_name = os.path.dirname(for_mx_dashboard_file)
    folder_name = os.path.basename(full_folder_name)
    print('1 folder_name=', folder_name)

    full_folder_name_up = os.path.dirname(full_folder_name)
    folder_up = os.path.basename(full_folder_name_up)
    print('folder_up=', folder_up)
    #print("DEBUG EXIT5")
    #exit(5)

    last_line_number_in_top_section, bottom_header = get_last_line_number_in_top_section(score_file)
    print('last_line_number_in_top_section=', last_line_number_in_top_section)

    if not last_line_number_in_top_section:
        log("ERROR parsing " + score_file)
        return

    print('bottom_header')
    print(bottom_header)



    if for_mx_dashboard_file:
        print('for_mx_dashboard_file=\n', for_mx_dashboard_file)
        df_for_mx_dashboard = pd.read_csv(for_mx_dashboard_file)  # used to join with bottom section of score file
        first_row = df_for_mx_dashboard.iloc[0].to_dict() # used to get date, test_loc, test_loc2, test_type, output_type
        first_row_date=first_row["date"]
        print('date=', date, type(date))
        print('first_row_date=', first_row_date, type(first_row_date), str(first_row_date))
        if not first_row_date or str(first_row_date)=='nan':
            print("No date in first_row")
        else:
            date=first_row_date
        


        test_loc=first_row["test_loc"]
        test_loc2=first_row["test_loc2"]
        test_type=first_row["test_type"]         
        output_type=first_row["output_type"]


        build_per_group = find_build_per_group(df_for_mx_dashboard, bottom_header)

        #df_top = read_top_section(score_file, last_line_number_in_top_section)
        df_top = read_top_section_with_group(score_file, last_line_number_in_top_section)

        df_top["date"]=date
        df_top["test_loc"]=test_loc
        df_top["test_loc2"]=test_loc2
        df_top["test_type"]=test_type
        df_top["output_type"]=output_type
        df_top["folder"]=folder_name
        df_top["folder_up"]=folder_up

        df_top['build'] = df_top['group_name'].map(build_per_group)
        print(df_top.info())
        print(df_top)
        #print("Exit before delete  ")
        #exit(0)
        
        table='kpi_score_gdc_summary'
        sql=f"DELETE from {table} WHERE date='{date}' and test_loc='{test_loc}' and folder='{folder_name}' and folder_up='{folder_up}'"
        print(sql)
        print("Before DELETE")
        db.dbExecute(sql)
        print("After DELETE")
        db.dbInsertDF(df_top, table)
        print("After INSERT to ", table)
        #print("Exit    ")
        #exit(0)

        #--------------------------------------------------  
        ### Process bottom section of score file (per dut)
        #-------------------------------------------------

        #df_bottom = read_bottom_section(score_file, last_line_number_in_top_section)
        df_bottom = read_bottom_section_with_group(score_file, last_line_number_in_top_section)
        print("df_bottom.info()=")
        print(df_bottom.info())
        print(df_bottom)

        
        cols=['score'] #,'val']
        dtypes = { col: df_bottom[col].dtype for col in cols }
        if all (dtype == 'object' for dtype in dtypes.values()):
          if df_bottom[cols].isna().any().any():
            print("Found None values")
            #print(df_top[cols])
            for col in cols:
              #print('col=', col)
              df_bottom[ col ] = df_bottom[col].fillna('')

          mask=pd.Series(True, index = df_bottom.index)
          for i, col in enumerate(cols):
            print(i, col, "build mask")
            mask &= ~df_bottom[col].str.contains('-', na=False)

          df_bottom = df_bottom[mask]    
          print(df_bottom[cols])
          df_bottom[cols] = df_bottom[cols].replace('',np.nan)
          df_bottom[cols] = df_bottom[cols].astype(float)


        df_bottom.loc[(df_bottom['score'] < 0.001) & (df_bottom['score'] > -0.001), 'score'] = 100.00 # WARNING  it can be -1
        print('df_bottom.info()=')
        print(df_bottom.info())


        print("process_gdc_kpi() \n score_file=", score_file)
        print("process_gdc_kpi() \n for_mx_dashboard_file=", for_mx_dashboard_file)

        print("\n\n DUT  FROM df_bottom")
        print("------------------------")
        score_duts = df_bottom['dut'].unique()
        score_duts.sort()
        for i, dut in enumerate(score_duts):
            print (i, dut)


        print("\n\n DUT_NUM FROM df_for_mx_dashboard")
        print("------------------------------------")
        for_mx_dashboard_duts = df_for_mx_dashboard['dut_num'].unique()
        for_mx_dashboard_duts.sort()
        for i, dut in enumerate(for_mx_dashboard_duts):
            print (i, dut)

        find_non_matching_duts(score_duts, for_mx_dashboard_duts, score_file, for_mx_dashboard_file)
        print(' before left_join_startswith len(df_bottom)=', len(df_bottom))
        #print(df_bottom)
        df_with_build = left_join_startswith(df_bottom, df_for_mx_dashboard)
        print(' after left_join_startswith  len(df_with_build)=', len(df_with_build))
        #print(df_with_build)
        #New columns: 'chipset', 'build_num', 'dut_model'
        # NOW we have 2 similar columns: dut (from score.csv file) and dut_num (this one we need to display)

        df_with_build = df_with_build.drop(['dut'], axis=1)
        df_with_build.rename(columns={
                'build_num' : 'build',
                'dut_model' : 'model',
                'dut_num' : 'dut',
                'group' : 'group_name'
            }, inplace=True)

        
        print('df_with_build.info()=')
        print(df_with_build.info())
        print(df_with_build)

        
        df_with_build["date"]=date
        df_with_build["folder"]=folder_name
        df_with_build["folder_up"]=folder_up
        df_with_build["test_loc"]=test_loc
        df_with_build["test_loc2"]=test_loc2
        df_with_build["test_type"]=test_type
        df_with_build["output_type"]=output_type
        
        column_order = ['date', 'test_loc', 'test_loc2', 'group_name',  'folder', 'folder_up',  'test_type', 'output_type', 'test', 'criteria', 'is_ref', 'dut', 'val', 'score', 'chipset', 'build', 'model']
        df_bottom_final = df_with_build[column_order]

        print('df_bottom_final.info()=')
        print(df_bottom_final.info())
        print(df_bottom_final)


        table='kpi_score_gdc'
        sql=f"DELETE from {table} WHERE date='{date}' and test_loc='{test_loc}' and folder='{folder_name}' and folder_up='{folder_up}'"
        print(sql)
        db.dbExecute(sql)
        print("before bottom insert")
        db.dbInsertDF(df_bottom_final, table)
        print("after bottom insert")



#-----------------------------------------------------
def populate_kpi_score_gdc(start_date, end_date=None, folder=None):
#-----------------------------------------------------
    print('populate_kpi_score_gdc() start_date=',start_date)
    #if True: #end_date:
    #    print('end_date=',end_date, type(end_date))
    dates = dates_in_range(start_date, end_date)
    #else:
    #    dates=[start_date]
    print("dates=", dates)
    #exit(0)
    
    for date in dates:
        print(date)
        #yyyymm=date[0:8]
        #folder=r"V:\FieldTestLogs\spotlight\2025-03\2025-03-12\SF"
        if not folder:
            folder=r"V:\FieldTestLogs\spotlight"+"\\"+date[0:7]+"\\"+date
        if not os.path.exists(folder):
            print("folder not exists", folder)
            continue

        print("folder=", folder)
        # under V:\FieldTestLogs\spotlight\2025-03\2025-03-11 find files for_mx_dashboard.csv
        # for every such file find _score.csv file 
        gdc_score_files= find_gdc_score_files(folder)
        print('gdc_score_files', gdc_score_files)
        for score_file, dashboard_file in gdc_score_files.items():
            print('score_file=',score_file)
            print('dashboard_file=', dashboard_file)
            print()
            process_gdc_kpi(date, score_file, dashboard_file)     

############################
if __name__ == "__main__":
############################

    args = parse_arguments()
    print('args.kpi_score=', args.kpi_score)
    POPULATE_KPI_SCORE = args.kpi_score
    POPULATE_KPI_SCORE_GDC = args.kpi_score_gdc

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
    if not os.path.exists(OUTPUT_DIR+"\\PDF"):
        os.makedirs(OUTPUT_DIR+"\\PDF")

    project=PROJECT.replace(' ','') + '_'
    output_excel_file = OUTPUT_DIR+"\\score_" + project + str(args.date) + ".xlsx"

    if args.date and not args.end_date and not args.back and not args.build:
        # process 1 day only: create MS Excel file or populate table kpi_score, do no send e-mail
        if POPULATE_KPI_SCORE_GDC:
            populate_kpi_score_gdc(str(args.date), str(args.date), args.folder)
            exit(0)

        all_dfs_per_location, all_dut_per_location = daily_score_report(args.date, output_excel_file)
        if not all_dfs_per_location and not all_dut_per_location:
            log("No input data for " +  str(args.date))
        else: 
            log('Report is created:' + output_excel_file)
            print('all_dfs_per_location =')
            for k, v in all_dfs_per_location.items():
              print('key=', k, 'number of records=', len(v))

    else:

        if args.date:
            start_date = str(args.date) #.strftime('%Y-%m-%d')
        elif args.back and not args.date:
            start_date = (datetime.now() - timedelta(days=args.back)).strftime('%Y-%m-%d')
            start_date = str(start_date)

        if POPULATE_KPI_SCORE_GDC:
            if args.end_date:
                end_date = str(args.end_date)
            else:
                end_date = str(args.date)   
            populate_kpi_score_gdc(start_date, end_date)
            exit(0)    

        # process several days, build trend report as pdf and send it as attachment
        # if GENERATE_ATTACHMENTS=False the following line will return None, None
        generated_pdfs , all_locations = trending_score_report(args)   #

        all_dfs_per_location, all_dut_per_location = daily_score_report(start_date, output_excel_file)
        if not all_dfs_per_location and not all_dut_per_location:
            log("No input data for " +  start_date)
            exit(1)

        if args.send_email : # and generated_pdfs:
            html_body = create_html(all_locations, all_dfs_per_location, all_dut_per_location, start_date)

            images_to_attach = get_images_for_email(all_locations)
            attachments = generated_pdfs + images_to_attach
            if os.path.exists(output_excel_file):
                attachments.append(output_excel_file)
            log("before send_mail " + str(start_date))
            send_mail(attachments, html_body)
            log("after send_mail")


