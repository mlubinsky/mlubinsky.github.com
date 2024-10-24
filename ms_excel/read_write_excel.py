import pandas as pd
import os
import re
from glob import glob
import openpyxl

summary_file=r"V:\RegressionLibrary\SPOT-ACE\kpi\PERFORMANCE_INDEX_REPORT_S23\performance_index_S24FE.xlsx"
#summary_file=r"V:\RegressionLibrary\SPOT-ACE\kpi\PERFORMANCE_INDEX_REPORT_S24\performance_index_S24FE.xlsx"
#summary_file=r"V:\RegressionLibrary\SPOT-ACE\kpi\PERFORMANCE_INDEX_REPORT_GS24\performance_index_GS24.xlsx"
if 'REPORT_S23' in summary_file:
     SHEET_NAME="S23"
elif 'REPORT_S24' in summary_file:
     SHEET_NAME="S24"
elif 'REPORT_GS24' in summary_file:
     SHEET_NAME="GS24"
else:
     print("cannot find sheet_name")
     exit(0)
#Location
#Metric
#Calculation
#Row Summary Fail Rate (# of fails/total number)
#------------------------------
def read_excel(summary_file):
#-------------------------------
  final_df=None
  df_all = pd.read_excel(summary_file, sheet_name=None)
  for sheet, df in df_all.items():
    print(sheet)
    df_50_95 = df[df['Metric'].str.contains('50%|95%', regex=True)]
    filtered_df = df_50_95[~df_50_95['Metric'].str.startswith('Along')]
    cols_of_interest=[]
    for i, col in enumerate(filtered_df.columns):
        if i < 4 and i !=2: # 2 is 'Calculation' column
                cols_of_interest.append(col)
                print(i, col)

    last_column = cols_of_interest[-1]

    df = filtered_df[cols_of_interest].copy()
    col_name = sheet + "\n" + " DUT fails / total"
    df.rename(columns={last_column: col_name}, inplace=True)
    print('\n\n relevant_columns')
    print(df)
    if final_df is None:

         final_df = df[['Location', 'Metric', col_name]].copy()
    else:
         final_df = pd.merge(final_df, df[['Location', 'Metric', col_name]],  on=['Location', 'Metric'], how='outer')

  print(final_df)
  return final_df

#------------------------
def write_df_to_file(df):
#-----------------------
  fname = SHEET_NAME+'.xlsx'
  output_writer = pd.ExcelWriter(fname, engine='openpyxl')
  df.to_excel(output_writer, sheet_name=SHEET_NAME, index=False)
  output_writer.save()
  print(fname)
  return fname

#-------------------------------
def process_ratio_string(value):
#-------------------------------
    # Remove parentheses and split the ratio into x and y
    if not value:
        return None
    if value == '100% Pass':
        return 0
    try:
        x, y = map(int, value.split('/'))
        return x / y
    except:
        return None
#----------------------
def add_summary_row(df):
#----------------------
    new_rows=[]
    for location, group in df.groupby('Location'):
        # For each group, calculate summary row
        summary_row = {'Location': 'Win Percentage'}
        for col in df.columns:
            if col in ['Location', 'Metric']:
              continue
             # Count rows where x/y < 0.5
            count_pass = group[col].apply(lambda x: process_ratio_string(x)).le(0.5).sum()   #lt

            count_total = group[col].count()
            print( 'count_total=',  count_total, 'count_pass =', count_pass)
            if count_total  > 0 and count_pass is not None:
              percent = count_pass * 100 / count_total
              percent = round(percent, 1)
              summary_row[col] = f'{percent}% ({count_pass} / {count_total})'
            else:
               summary_row[col] = None

        # Add 'Summary' to all other columns
        summary_row.update({col: None for col in df.columns if col not in summary_row})

        # Append the group and its summary row
        new_rows.append(group)
        new_rows.append(pd.DataFrame([summary_row]))

    # Concatenate the original dataframe with the summary rows
    result_df = pd.concat(new_rows).reset_index(drop=True)
    return result_df

from openpyxl.utils import get_column_letter
#---------------------
def add_colors(fname):
#---------------------
  wb = openpyxl.load_workbook(fname)
  ws = wb[SHEET_NAME]
  yellow_fill = openpyxl.styles.fills.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
  blue_fill   = openpyxl.styles.fills.PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
  green_fill  = openpyxl.styles.fills.PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
  pink_fill   = openpyxl.styles.fills.PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
  light_green_fill  = openpyxl.styles.fills.PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
  pink_fill         = openpyxl.styles.fills.PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

  for cell in ws[1]:  # Apply color to header  in the first row
     cell.fill = blue_fill
     cell.alignment = cell.alignment.copy(wrapText=True)

  current_column_A = ws.cell(row = 2, column=1).value
  max_row_count = ws.max_row
  max_col_count = ws.max_column
  use_color_for_A=False
  for row in range(2, max_row_count+1):
    column_A = ws.cell(row = row , column=1).value

    if False:
     column_A = ws.cell(row = row , column=1).value
     if  column_A != current_column_A:
       current_column_A = column_A
       use_color_for_A = not use_color_for_A
     if use_color_for_A:
        ws.cell(row = row, column = 1).fill = blue_fill
        ws.cell(row = row, column = 2).fill = blue_fill

    for col in range(3, max_col_count+1):
        val =  ws.cell(row = row , column=col).value
        if val=="100% Pass":
            ws.cell(row = row , column=col).fill = light_green_fill
        elif val:
            tokens = val.split("/")
            if len(tokens) == 2:
              try:
                fail=int(tokens[0])
                total=int(tokens[1])
                if fail*2 <= total:
                    ws.cell(row = row , column=col).fill = light_green_fill
              except:
                pass

  for row in range(2, max_row_count+1):
      column_A = ws.cell(row = row , column=1).value
      if column_A == 'Win Percentage':
         for col in range(1, max_col_count+1):
             ws.cell(row = row , column=col).fill = blue_fill


  for i, column_cells in enumerate(ws.columns):

      if i < 2:
        ws.column_dimensions[get_column_letter(i+1)].width = 25
      else:
         ws.column_dimensions[get_column_letter(i+1)].width = 15

  wb.save(fname)
  return True


####   MAIN

df = read_excel(summary_file)
df = add_summary_row(df)
fname = write_df_to_file(df)
add_colors(fname)
