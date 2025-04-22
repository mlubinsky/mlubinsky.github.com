import os
import sys
import argparse
import random
import logging

from datetime import datetime, timedelta
from collections import defaultdict
from collections import Counter
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font,  PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from shutil import rmtree

C_DIR=r"C:\MY_DIR"
print(C_DIR)
sys.path.append(C_DIR)

# Set up the logger
log_file="my.log"
logging.basicConfig(
    level=logging.INFO,  # Log level
    format='%(asctime)s - %(message)s',  # Log format
    handlers=[
        logging.FileHandler(log_file),  # Log to a file
        logging.StreamHandler(sys.stdout)    # Log to the console
    ]
)

# Wrapper function for logging 
def log(message, level='info'):
    logging.info(message)

#plt.set_loglevel('WARNING')

from utils.DBHelper import DBHelper
from spotemail import SpotEmailNew

db = DBHelper(host='xxx.yyy.zz.dd') 

OUTPUT_DIR=r"YYY"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

IMG_DIR=os.path.join(SCRIPT_DIR, "images")
if os.path.exists(IMG_DIR):
       rmtree(IMG_DIR, ignore_errors=True)

DEBUG=False
THRESHOLD=90.0
#------------------
def dates_in_range(start, end):
#-----------------
    # Convert the input strings to datetime objects
    start_date = datetime.strptime(start, '%Y-%m-%d')
    end_date = datetime.strptime(end, '%Y-%m-%d')

    # Create an empty list to store the dates
    date_list = []

    # Generate the list of dates between start and end, excluding the end date
    current_date = start_date #+ timedelta(days=1)
    while current_date <= end_date:
        date_list.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)

    return date_list

#---------------------------
def valid_date(date_str):
#---------------------------
    """Validate the date format is YYYY-MM-DD."""
    #log("Validate input date: "+ date_str)
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        log("Error in input date: "+ date_str)
        raise argparse.ArgumentTypeError(f"Invalid date format: '{date_str}'. Expected format: YYYY-MM-DD.")
#----------------------
def parse_arguments():
#---------------------
    parser = argparse.ArgumentParser(description="Fetch records from PostgreSQL table T.")

    # Adding command line arguments
    parser.add_argument('--date', type=valid_date, help='Start date in format YYYY-MM-DD.')
    parser.add_argument('--end_date', type=valid_date, help='End date in format YYYY-MM-DD.')
    parser.add_argument('--back', type=int, help='Number of days back from --date.')
    parser.add_argument('--send_email', action='store_true', help='Send e-mail if this flag is provided')
    args = parser.parse_args()


    # Validations
    if args.date and args.build:
        log("Error: --date and --build cannot be provided together.")
        sys.exit(1)

    if args.back and args.end_date:
        log("Error: --back and --end_date cannot be provided together.")
        sys.exit(1)

    return args
#-----------------------------------
def send_mail(list_of_attachments, title, body, html_body=None):
#-----------------------------------
    print(" send_mail() start")
    RECIPIENTS_DEVELOPER_MODE = 'mlubinsky@hotmail.com'

    print('list_of_attachments=')
    print(list_of_attachments)

    if list_of_attachments:
        attachments = ','.join(list_of_attachments)
        #body="MX score trend reports are attached."
    else:
        attachments = ''
        #body="No attachments."

    print('body=', body)

    email_engine=SpotEmailNew(sender=None, recipients=RECIPIENTS_DEVELOPER_MODE)
    print(" send_mail() before")
    email_engine.send_email(subject=title,
                            text=body,
                            files=attachments,
                            html=html_body
                            )
    print(" send_mail() end")

#---------------------------------------
def gdc_builds(location, start_date, end_date, not_is_ref_only=True):
#---------------------------------------
    test_loc = location[0]
    sql=f"""SELECT distinct TO_CHAR(date,'YYYY-MM-DD') as date, build 
            FROM summary_table
            WHERE date >='{start_date}'  AND date <='{end_date}'  
            AND test_loc ='{test_loc}'
            AND NOT is_ref
            and (folder_up like '%DEV%' or test_loc not like 'DSK%')
            and folder like '%Flip%'
            ORDER BY date, build"""

    print(sql)   
    dates_and_builds = db.dbFetchAll(sql)

    for i, rec in enumerate(dates_and_builds): 
        print(i, rec)
    return dates_and_builds


#---------------------------------------
def gdc_scores(location, dates):
#---------------------------------------
# Report only REF_B6  (by Balaji request)

    print("gdc_scores() test_loc =", location[0] )
    test_loc = location[0]
    print ('set(dates)=')
    print (set(dates))
     

    set_dates = str(set(dates))[1:-1]
    print('set_dates=')
    print(set_dates)

    sql=f"""select group_name,  test,   is_ref, build, TO_CHAR(date,'YYYY-MM-DD') as date , score   
    from summary_table
    where test_loc='{test_loc}' and criteria= '_Average'
    and date in ({set_dates})
    and (folder_up like '%DEV%' or test_loc not like 'DSK%')
    and folder like '%Flip%'
    and (group_name  like 'REF%B6%' or not is_ref)
    order by test, is_ref DESC , group_name, date, build"""

    print(sql)   
    scores = db.dbFetchAll(sql)

    if not scores:
        print ("no scores", test_loc)
        return scores
  

    for i, rec in enumerate(scores): 
        print(i, rec)
    return scores


#---------------------------------------
def gdc_locations(start_date, end_date):
#---------------------------------------
    print("gdc_locations()", start_date, end_date)
    sql=f"""SELECT distinct test_loc, test_loc2, test_type 
        FROM summary_table 
        WHERE date >='{start_date}'  AND date <='{end_date}'  
        and (folder_up like '%DEV%' or test_loc not like 'DSK%')
        and folder like '%Flip%'
        and not is_ref
        ORDER BY test_loc"""

    print(sql)   
    test_locations = db.dbFetchAll(sql)

    print('locations are:')
    for i, rec in enumerate(test_locations): 
        print(i, rec)

    return  test_locations   

#--------------------------------------------------
def gdc_sheet(ws, location, start_date, end_date):
#-------------------------------------------------

    dates_and_builds = gdc_builds(location, start_date, end_date)
    if not dates_and_builds: # no data
        print(" no data for ", location)
        return

    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    blue_fill   = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    green_fill  = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    pink_fill   = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    grey_fill =  PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')

    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)

    font_bold = Font(bold=True, size=12)

   # make wide
    ws.column_dimensions['A'].width = 40

    A1= ' '.join(location) 

    dates= [A1]
    builds=['Build ->']
    for el in dates_and_builds:
       dates.append(el[0])
       builds.append(el[1])  


    ws.append(dates)  # row 1
    ws.append(builds) # row 2 

    ws.cell(row = 1, column = 1).fill = yellow_fill
    ws.cell(row = 1, column = 1).font = font_bold

    for col in range(2, ws.max_column+1):
       ws.cell(row = 2, column = col).fill = blue_fill 
    
    current_row = 2
    previous_group_and_test=None  # add line with scores to MS Exclel based on this 
    previous_test = None   # add empty line between tests

    scores = gdc_scores(location, dates[1:])
    for i, rec in enumerate(scores): 
        #print(i, rec)
        group_name=rec[0]  
        test=rec[1]  
        is_ref=rec[2]
        idx = test.find('.')
        if idx > 0:
            test=test[idx+1:]

        if (i == 0) : #or (test != previous_test):  # add new line 
                print('score_rec =',i, 'group=', group_name, 'test=', test, 'previous_test=',previous_test)

                previous_test=test 
                ws.append([test])
                current_row +=1
                ws.cell(row = current_row, column = 1).font = font_bold    
                for col in range(1, ws.max_column+1):
                    ws.cell(row = current_row, column = col).fill = green_fill

        group_and_test =  group_name + "_" + test
        if group_and_test != previous_group_and_test: # add new line to MS Excel
            if previous_group_and_test:
                ws.append(line)
                current_row +=1


                if  (test != previous_test):  # add new line 
                  print('score_rec =',i, 'group=', group_name, 'test=', test, 'previous_test=',previous_test)

                  previous_test=test 
                  ws.append([test])
                  current_row +=1
                  ws.cell(row = current_row, column = 1).font = font_bold    
                  for col in range(1, ws.max_column+1):
                    ws.cell(row = current_row, column = col).fill = green_fill

            
            line = [group_and_test]+['']*(len(dates)-1)  
            previous_group_and_test = group_and_test


        build=rec[3] 
        date=rec[4]  
        score=rec[5]


        # find build and date index in build dates
        if not is_ref:
          index = dates_and_builds.index((date,build))
          #print("index=", index)
          line[index+1]=score
        else:
            #print("match only date", date)
            for idx, d in enumerate(dates): # 1st element in dates is '' 
                if d == date:
                    line[idx]=score
      
    for col in range(2, ws.max_column+1): 
        letter =  get_column_letter(col) 
        #print('letter=', letter)
        ws.column_dimensions[letter].width = 20



    for row_number , row in enumerate(ws, start=1):
        for cell in row:
          cell.alignment = Alignment(horizontal='left', vertical='center')

          if row_number > 2 and cell.column > 1 and cell.data_type =='n':
            if cell.value and cell.value < 90.0:
              cell.fill = pink_fill

          cell.border = border
        #if row_number == 3:
        #      exit(0)

#---------------------------------
def main(star_date, end_date, email):
#--------------------------------
    print(start_date, end_date)

    test_locations=gdc_locations(start_date, end_date)
    if not test_locations:
        return None
    # get data from kpi_score_gdc_summary for the date range
    # one location  - one MS Excel sheet
    OUTPUT_DIR=r"YYY"
    MS_Excel_file=f"""trend_{start_date}_{end_date}.xlsx"""
    excel_file = os.path.join(OUTPUT_DIR, MS_Excel_file)
    print(excel_file)
   
    wb = Workbook()
    ws = wb.active
    for location in test_locations:
        sheet=location[0]
        ws = wb.create_sheet(title=sheet)
        gdc_sheet(ws, location, start_date, end_date)
        #break

    del wb['Sheet']

    print(excel_file)
    wb.save(excel_file)
    wb.close()

    if email:
        attachments=[]
        title = "Flip build trend MS Excel is attached."
        body = title 
        if os.path.exists(excel_file):
                attachments.append(excel_file)
        send_mail(attachments, title, body)
     

############################
if __name__ == "__main__":
############################

    args = parse_arguments()
    if args.date:
            start_date = str(args.date) #.strftime('%Y-%m-%d')
    elif args.back and not args.date:
            start_date = (datetime.now() - timedelta(days=args.back)).strftime('%Y-%m-%d')
            start_date = str(start_date)
    else:
        print("provide --start_date or --back")
        exit(0)

    if args.end_date:
        end_date = str(args.end_date)
    else:
        end_date = (datetime.now()).strftime('%Y-%m-%d')

    main(start_date, end_date, args.send_email)

    exit(0)

